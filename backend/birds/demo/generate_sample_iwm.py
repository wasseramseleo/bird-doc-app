#!/usr/bin/env python3
"""Generate a *synthetic* IWM-format capture sheet standing in for the real
export we'll receive later (Phase 1 of the Referenzprojekt — see ADR 0012).

This is NOT real data and contains no personal data — it is invented field data
for a plausible reed-bed ringing operation at Illmitz / Neusiedlersee
(Burgenland), caught with Japannetz (Fangmethode ``M``) across two summer+autumn
seasons. It mirrors the *authentic* IWM ``Datenmeldung`` conventions seen in the
export template's own example rows — notably Geschlecht as letters ``U/M/W`` (not
the integers our own ``iwm_export.py`` writes), integer Alter, string category
codes, ``Zusatzmarkierung="ZZ"`` — so it exercises the real import path.

It deliberately looks like *pre-anonymisation* input: several real-ish Beringer
and sub-sites, sequential ring runs, Wiederfänge that reuse an Erstfang's ring
identity, and two Sonderarten. The anonymiser (ADR 0012) later collapses those
onto the curated demo cast and shifts the years.

Deterministic: fixed seed ⇒ byte-stable output. Regenerate with::

    backend/.venv/bin/python backend/birds/demo/generate_sample_iwm.py
"""

from __future__ import annotations

import argparse
import datetime as dt
import random
from pathlib import Path

import openpyxl

# Authentic IWM "Fangdaten" header row, in template column order (1..32).
HEADERS = [
    "Ring",
    "Ringnummer",
    "Ringstatus",
    "Zusatzmarkierung",
    "Art",
    "Fangmethode",
    "Lockmittel",
    "Geschlecht",
    "Alter",
    "Datum",
    "Uhrzeit",
    "Ortskodierung",
    "Geo-Koordinaten",
    "Zustand",
    "Umstand",
    "Flügellänge",
    "Teilfederlänge",
    "Gewicht",
    "Tarsus",
    "Fett",
    "Muskel",
    "Intensität",
    "Fortschritt",
    "Handschwingen",
    "Brutfleck",
    "Kloake",
    "Netz",
    "Ort",
    "Region",
    "Land",
    "Bemerkungen",
    "BeringerIn",
]

SCHEME_CODE = "AUW"
FANGMETHODE = "M"  # Japannetz
LOCKMITTEL = "N"  # sicher kein Lockmittel
UMSTAND = "20"  # wissenschaftliche Beringung (as in the template examples)
ZUSTAND = "8"  # lebend, unverletzt freigelassen
REGION = "Burgenland"
LAND = "Austria"

# Real-ish ringers (collapsed to the demo cast later). Kürzel = first letter of
# first name + first two of surname (Austrian standard).
RINGERS = [
    ("Johanna Gruber", "JGR"),
    ("Matthias Wallner", "MWA"),
    ("Sophie Bauer", "SBA"),
    ("Lukas Steiner", "LST"),
    ("Elena Hofer", "EHO"),
]
# Two dominant ringers (a real roster is rarely uniform).
RINGER_WEIGHTS = [30, 26, 18, 14, 12]

# Illmitz / Neusiedlersee sub-sites. name, Ortskodierung, lat, lon, #nets.
STATIONS = [
    ("Illmitz Zicklacke", "NS01", 47.769000, 16.792000, 8),
    ("Illmitz Sandeck", "NS02", 47.742000, 16.830000, 10),
    ("Illmitz Schilfgürtel", "NS03", 47.756000, 16.808000, 12),
]

# Ring size series (Austrian scheme). size -> running counter (last consumed).
RING_BASE = {"V": 18000, "S": 5400, "T": 2100, "X": 31000, "P": 900}

# species: de, sci, size, wing(lo,hi), weight(lo,hi), tarsus(lo,hi), freq,
# dimorphic, season affinity ('both'|'summer'|'autumn').
SPECIES = [
    (
        "Teichrohrsänger",
        "Acrocephalus scirpaceus",
        "V",
        (62, 69),
        (10.5, 14.5),
        (21, 24),
        22,
        False,
        "both",
    ),
    (
        "Schilfrohrsänger",
        "Acrocephalus schoenobaenus",
        "V",
        (62, 69),
        (10.0, 14.0),
        (20, 23),
        16,
        False,
        "both",
    ),
    (
        "Sumpfrohrsänger",
        "Acrocephalus palustris",
        "V",
        (62, 69),
        (10.5, 14.0),
        (21, 24),
        8,
        False,
        "both",
    ),
    (
        "Drosselrohrsänger",
        "Acrocephalus arundinaceus",
        "S",
        (90, 101),
        (25, 35),
        (32, 37),
        3,
        False,
        "summer",
    ),
    (
        "Mariskensänger",
        "Acrocephalus melanopogon",
        "V",
        (53, 60),
        (9.5, 12.5),
        (20, 23),
        2,
        False,
        "both",
    ),
    (
        "Rohrschwirl",
        "Locustella luscinioides",
        "V",
        (66, 73),
        (13, 18),
        (22, 25),
        4,
        False,
        "summer",
    ),
    ("Feldschwirl", "Locustella naevia", "X", (60, 67), (11, 15), (20, 23), 2, False, "autumn"),
    ("Mönchsgrasmücke", "Sylvia atricapilla", "V", (70, 78), (15, 22), (19, 22), 14, True, "both"),
    ("Gartengrasmücke", "Sylvia borin", "V", (76, 83), (15, 22), (19, 22), 5, False, "autumn"),
    ("Dorngrasmücke", "Sylvia communis", "V", (66, 73), (12, 18), (21, 24), 6, True, "both"),
    ("Klappergrasmücke", "Sylvia curruca", "X", (60, 66), (10, 14), (19, 22), 4, False, "autumn"),
    ("Zilpzalp", "Phylloscopus collybita", "X", (56, 63), (6, 9.5), (18, 21), 8, False, "autumn"),
    ("Fitis", "Phylloscopus trochilus", "X", (64, 71), (7, 11), (18, 21), 6, False, "autumn"),
    ("Rotkehlchen", "Erithacus rubecula", "V", (70, 76), (15, 20), (24, 27), 7, False, "autumn"),
    ("Blaukehlchen", "Luscinia svecica", "V", (70, 78), (15, 21), (26, 29), 3, True, "both"),
    ("Nachtigall", "Luscinia megarhynchos", "V", (82, 90), (18, 27), (26, 29), 2, False, "summer"),
    ("Rohrammer", "Emberiza schoeniclus", "V", (72, 80), (15, 21), (19, 22), 9, True, "both"),
    ("Bartmeise", "Panurus biarmicus", "V", (55, 63), (12, 16), (20, 23), 5, True, "both"),
    ("Beutelmeise", "Remiz pendulinus", "X", (52, 58), (8, 11), (15, 18), 2, False, "autumn"),
    ("Blaumeise", "Cyanistes caeruleus", "V", (62, 68), (9, 12.5), (16, 18), 5, False, "both"),
    ("Kohlmeise", "Parus major", "V", (72, 80), (15, 21), (19, 22), 4, False, "both"),
    ("Zaunkönig", "Troglodytes troglodytes", "X", (44, 50), (8, 11), (17, 19), 3, False, "autumn"),
    ("Seidensänger", "Cettia cetti", "V", (55, 63), (10, 15), (22, 25), 2, False, "both"),
    ("Rauchschwalbe", "Hirundo rustica", "V", (118, 135), (15, 22), (10, 12), 6, False, "autumn"),
    ("Amsel", "Turdus merula", "T", (118, 135), (85, 110), (33, 37), 3, True, "both"),
    ("Star", "Sturnus vulgaris", "P", (122, 135), (68, 92), (27, 30), 2, False, "both"),
    ("Eisvogel", "Alcedo atthis", "S", (72, 80), (34, 46), (9, 11), 1, False, "both"),
    ("Neuntöter", "Lanius collurio", "V", (88, 97), (25, 34), (22, 25), 2, True, "summer"),
]

BEMERKUNGEN_POOL = [
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",  # mostly blank
    "Zecken am Kopf",
    "leichte Handschwingenmauser",
    "Schwanz frisch vermausert",
    "Lausfliege",
    "Gefieder stark abgenutzt",
    "Fettdepot deutlich sichtbar",
]


def _date_pool(years):
    """Weighted date pool: autumn-migration heavy, summer breeding lighter."""
    pool = []
    for y in years:
        # Summer breeding window (lighter weight).
        d = dt.date(y, 6, 5)
        while d <= dt.date(y, 7, 25):
            pool += [d] * 1
            d += dt.timedelta(days=1)
        # Autumn migration window, peaking late Aug–Sep (heavier weight).
        d = dt.date(y, 8, 1)
        while d <= dt.date(y, 10, 25):
            peak = 5 if dt.date(y, 8, 20) <= d <= dt.date(y, 9, 25) else 3
            pool += [d] * peak
            d += dt.timedelta(days=1)
    return pool


def _season_of(d: dt.date) -> str:
    return "summer" if d.month in (6, 7) else "autumn"


def _pick_species(rng, season):
    cands = [s for s in SPECIES if s[8] in ("both", season)]
    weights = [s[6] for s in cands]
    return rng.choices(cands, weights=weights, k=1)[0]


def _pick_time(rng, species_de) -> dt.time:
    if species_de == "Rauchschwalbe":  # reed roost, evening
        h = rng.choice([18, 19])
        return dt.time(h, rng.choice([0, 15, 30, 45]))
    total = rng.randint(5 * 60 + 30, 10 * 60)  # 05:30–10:00
    return dt.time(total // 60, (total % 60) // 5 * 5)


def _measure(rng, sp, age, sex, season):
    de, sci, size, wing, wt, tar, freq, dim, seas = sp
    row = {}
    row["Flügellänge"] = rng.randint(int(wing[0]), int(wing[1]))
    row["Gewicht"] = round(rng.uniform(wt[0], wt[1]), 1)
    row["Tarsus"] = round(rng.uniform(tar[0], tar[1]), 1) if rng.random() < 0.6 else None
    row["Teilfederlänge"] = (
        rng.randint(int(wing[0] * 0.72), int(wing[0] * 0.82)) if rng.random() < 0.15 else None
    )
    # Fett: autumn migrants carry more.
    if season == "autumn":
        row["Fett"] = str(rng.choices([1, 2, 3, 4, 5], weights=[15, 30, 30, 18, 7])[0])
    else:
        row["Fett"] = str(rng.choices([0, 1, 2, 3], weights=[30, 40, 22, 8])[0])
    row["Muskel"] = str(rng.choices([1, 2, 3], weights=[18, 66, 16])[0])
    row["Intensität"] = str(rng.choice([0, 1, 2])) if rng.random() < 0.75 else None
    # Fortschritt (Kleingefieder) only for diesjährige birds (age 3).
    row["Fortschritt"] = (
        rng.choices(["J", "U", "M", "N"], weights=[10, 30, 40, 20])[0]
        if age == 3 and rng.random() < 0.7
        else None
    )
    if rng.random() < 0.55:
        row["Handschwingen"] = str(
            rng.choices([0, 1, 2, 4], weights=[8, 60, 18, 14])[0]
            if season == "summer"
            else rng.choices([0, 1], weights=[20, 80])[0]
        )
    else:
        row["Handschwingen"] = None
    return row


def _sex(rng, dimorphic):
    if dimorphic:
        return rng.choices(["U", "M", "W"], weights=[40, 30, 30])[0]
    return rng.choices(["U", "M", "W"], weights=[80, 10, 10])[0]


def _age(rng, season):
    if season == "summer":
        return rng.choices([1, 3, 4, 6], weights=[15, 20, 45, 20])[0]
    return rng.choices([2, 3, 4], weights=[10, 72, 18])[0]


def build_rows(rng, n_erstfang, n_wiederfang):
    rings = dict(RING_BASE)
    ringed = []  # registry of ringed birds for recaptures
    rows = []
    dates = _date_pool([2023, 2024])

    def base_ctx(d, station):
        name, code, lat, lon, nets = station
        return {
            "Ring": SCHEME_CODE,
            "Zusatzmarkierung": "ZZ",
            "Fangmethode": FANGMETHODE,
            "Lockmittel": LOCKMITTEL,
            "Umstand": UMSTAND,
            "Zustand": ZUSTAND,
            "Datum": dt.datetime(d.year, d.month, d.day),
            "Ortskodierung": code,
            "Geo-Koordinaten": f"{lat:.6f}, {lon:.6f}",
            "Netz": str(rng.randint(1, nets)),
            "Ort": name,
            "Region": REGION,
            "Land": LAND,
        }

    # --- Erstfänge ---
    for _ in range(n_erstfang):
        d = rng.choice(dates)
        season = _season_of(d)
        sp = _pick_species(rng, season)
        de, sci, size, wing, wt, tar, freq, dim, seas = sp
        station = rng.choice(STATIONS)
        ringer = rng.choices(RINGERS, weights=RINGER_WEIGHTS, k=1)[0]
        rings[size] += 1
        ringnr = f"{size}{str(rings[size]).zfill(5)}"
        age = _age(rng, season)
        sex = _sex(rng, dim)
        row = base_ctx(d, station)
        row.update(
            {
                "Ringnummer": ringnr,
                "Ringstatus": "E",
                "Art": de,
                "Geschlecht": sex,
                "Alter": age,
                "Uhrzeit": _pick_time(rng, de),
                "Bemerkungen": rng.choice(BEMERKUNGEN_POOL) or None,
                "BeringerIn": ringer[1],
            }
        )
        row.update(_measure(rng, sp, age, sex, season))
        rows.append(row)
        ringed.append({"ringnr": ringnr, "size": size, "sp": sp, "sex": sex, "date": d})

    # --- Wiederfänge: reuse an earlier bird's ring identity, later date ---
    later = sorted(dates)
    for _ in range(n_wiederfang):
        bird = rng.choice(ringed)
        after = [x for x in later if x > bird["date"]]
        if not after:
            continue
        d = rng.choice(after)
        season = _season_of(d)
        sp = bird["sp"]
        de = sp[0]
        station = rng.choice(STATIONS)
        ringer = rng.choices(RINGERS, weights=RINGER_WEIGHTS, k=1)[0]
        # An older bird: if recaught a season later it is at least vorjährig.
        age = 4 if d.year == bird["date"].year else rng.choice([5, 6])
        row = base_ctx(d, station)
        row.update(
            {
                "Ringnummer": bird["ringnr"],
                "Ringstatus": "W",
                "Art": de,
                "Geschlecht": bird["sex"],
                "Alter": age,
                "Uhrzeit": _pick_time(rng, de),
                "Bemerkungen": "Wiederfang",
                "BeringerIn": ringer[1],
            }
        )
        row.update(_measure(rng, sp, age, bird["sex"], season))
        row["Fortschritt"] = None  # not a diesjährig bird any more
        rows.append(row)

    # --- Sonderarten (exercise the importer's special-kind path) ---
    d = dt.date(2024, 9, 12)
    st = STATIONS[0]
    r = base_ctx(d, st)
    rings["V"] += 1
    r.update(
        {
            "Ringnummer": f"V{str(rings['V']).zfill(5)}",
            "Ringstatus": "E",
            "Art": "Art nicht in der Liste (Aves ignota)",
            "Geschlecht": "U",
            "Alter": 3,
            "Uhrzeit": dt.time(7, 20),
            "Flügellänge": 66,
            "Gewicht": 12.4,
            "Tarsus": 22.5,
            "Fett": "2",
            "Muskel": "2",
            "Intensität": "1",
            "Handschwingen": "1",
            "Bemerkungen": (
                "Unbestimmter Acrocephalus, Verdacht Buschrohrsänger; "
                "Fotos und Vermessung an Vogelwarte übermittelt"
            ),
            "BeringerIn": "JGR",
        }
    )
    rows.append(r)

    d = dt.date(2023, 9, 3)
    r = base_ctx(d, STATIONS[1])
    rings["X"] += 1
    r.update(
        {
            "Ringnummer": f"X{str(rings['X']).zfill(5)}",
            "Ringstatus": "E",
            "Art": "Art nicht in der Liste (Aves ignota)",
            "Geschlecht": "U",
            "Alter": 3,
            "Uhrzeit": dt.time(8, 5),
            "Flügellänge": 61,
            "Gewicht": 8.2,
            "Tarsus": 19.8,
            "Fett": "3",
            "Muskel": "2",
            "Intensität": "1",
            "Handschwingen": "1",
            "Bemerkungen": "Kleiner Laubsänger, Bestimmung unsicher; Rufaufnahme gesichert",
            "BeringerIn": "MWA",
        }
    )
    rows.append(r)

    # Ring Vernichtet: a ring taken out of service — no bird data.
    d = dt.date(2024, 8, 28)
    r = base_ctx(d, STATIONS[2])
    rings["V"] += 1
    r.update(
        {
            "Ringnummer": f"V{str(rings['V']).zfill(5)}",
            "Ringstatus": None,
            "Art": "Ring Vernichtet",
            "Geschlecht": None,
            "Alter": None,
            "Uhrzeit": dt.time(6, 40),
            "Bemerkungen": "Ring beim Anlegen deformiert, vernichtet",
            "BeringerIn": "SBA",
        }
    )
    rows.append(r)

    rows.sort(key=lambda x: (x["Datum"], x.get("Uhrzeit") or dt.time(0, 0)))
    return rows


def write_workbook(rows, out_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fangdaten"
    ws.append(HEADERS)
    col = {h: i + 1 for i, h in enumerate(HEADERS)}
    for r_i, row in enumerate(rows, start=2):
        for h, c in col.items():
            v = row.get(h)
            if v is not None:
                ws.cell(row=r_i, column=c, value=v)
        ws.cell(row=r_i, column=col["Datum"]).number_format = "DD.MM.YYYY"
        ws.cell(row=r_i, column=col["Uhrzeit"]).number_format = "HH:MM"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "-o", "--out", type=Path, default=Path(__file__).with_name("sample_iwm_illmitz.xlsx")
    )
    p.add_argument("--erstfaenge", type=int, default=300)
    p.add_argument("--wiederfaenge", type=int, default=40)
    p.add_argument("--seed", type=int, default=42)
    a = p.parse_args()
    rng = random.Random(a.seed)
    rows = build_rows(rng, a.erstfaenge, a.wiederfaenge)
    write_workbook(rows, a.out)
    n_w = sum(1 for r in rows if r["Ringstatus"] == "W")
    n_son = sum(
        1 for r in rows if r["Art"] in ("Ring Vernichtet", "Art nicht in der Liste (Aves ignota)")
    )
    print(f"Wrote {len(rows)} rows to {a.out}")
    print(f"  Erstfänge: {len(rows) - n_w - n_son} · Wiederfänge: {n_w} · Sonderarten: {n_son}")
    arts = {}
    for r in rows:
        arts[r["Art"]] = arts.get(r["Art"], 0) + 1
    print(
        "  Arten:",
        ", ".join(f"{k}×{v}" for k, v in sorted(arts.items(), key=lambda x: -x[1])[:8]),
        "…",
    )


if __name__ == "__main__":
    main()
