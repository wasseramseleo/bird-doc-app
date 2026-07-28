from io import BytesIO
from pathlib import Path

import openpyxl
from django.utils.timezone import localtime
from openpyxl.styles import PatternFill

from .models import DataEntry, Species

TEMPLATE_PATH = (
    Path(__file__).resolve().parent / "templates" / "iwm" / "Datenmeldung_Vorlage_IWM.xlsx"
)
SHEET_NAME = "Fangdaten"

# Nicht-Standard-Fang (ADR 0026): its row is background-filled so the Beringer can
# spot it at a glance on their own review — purely visual, since the Meldestelle
# ignores cell formatting. A soft amber, applied across the whole data row.
NON_STANDARD_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
# A no-fill reset, applied to every data cell before writing so a template example
# row's leftover styling never leaks onto a real export row.
_NO_FILL = PatternFill(fill_type=None)
# What a Nicht-Standard-Fang blanks, phrased over the *source* of the value rather
# than as a fixed column list (ADR 0034): the method columns the **Projekt**
# supplies. The capture happened outside the Projekt's standard protocol, so those
# defaults do not describe it (ADR 0026). Fangmethode and Lockmittel are always the
# Projekt's, so they always blank; Umstand is the Projekt's too — unless the capture
# is also a Tot-Fund, whose 08 is a fact about *that* capture and therefore survives
# the combination. Zustand is never Projekt-derived and never blanks.
_ALWAYS_PROJECT_SOURCED_COLUMNS = frozenset({"Fangmethode", "Lockmittel"})


def _project_sourced_columns(entry):
    """The columns whose value this capture takes from its Projekt — exactly the
    ones a Nicht-Standard-Fang leaves empty."""
    if entry.is_dead_recovery:
        return _ALWAYS_PROJECT_SOURCED_COLUMNS
    return _ALWAYS_PROJECT_SOURCED_COLUMNS | {"Umstand"}


# The Tot-Fund's own codes (ADR 0034). Both are *derived* from the Fangmarker at
# export time and never stored on the capture: the Beringer picks no code.
TOT_FUND_UMSTAND = "08"
TOT_FUND_ZUSTAND = "2"
# "lebend, unverletzt freigelassen" — what every other row asserts. The authentic
# Datenmeldung has no blank Zustand cell, so BirdDoc emits one on every row; see
# the ADR's Consequences for the imprecision this knowingly accepts.
DEFAULT_ZUSTAND = "8"

# "No additional marking" — every authentic Datenmeldung row carries it (the
# importer reads and discards it, having no model field to land it in).
ZUSATZMARKIERUNG = "ZZ"

# The model stores Geschlecht as an integer (0/1/2); the authentic sheet carries
# it as a letter (U/M/W). This mapping is the inverse of the importer's, so
# export and import round-trip.
_SEX_TO_CODE = {
    DataEntry.Sex.UNKNOWN: "U",
    DataEntry.Sex.MALE: "M",
    DataEntry.Sex.FEMALE: "W",
}


def _sex_code(entry):
    """The authentic Geschlecht letter for a capture, or ``None`` when unset (a
    Sonderart row carries no sex)."""
    return _SEX_TO_CODE.get(entry.sex)


def _text_code(value):
    """An authentic category code is text in the sheet; write the integer the
    model stores as a string, leaving a blank cell blank."""
    return None if value is None else str(value)


def _geo_coordinates(entry):
    """Format the station's coordinates as ``lat, lon`` decimal degrees (dot separator)."""
    station = entry.ringing_station
    if not station or station.latitude is None or station.longitude is None:
        return None
    return f"{station.latitude}, {station.longitude}"


def _breeding_flag(value):
    """A breeding-indicator column (Brutfleck, Kloake) carries "J" when the flag
    is set and stays blank otherwise (issue #375)."""
    return "J" if value else None


# Parasit (ADR 0027): code → label for the fixed, app-wide vocabulary. The IWM
# template has no Parasit column, so each selected type's label goes into the
# Bemerkungen column (exactly where Milben used to land).
_PARASIT_LABELS = {code.value: str(code.label) for code in DataEntry.Parasit}


def _parasite_labels(entry):
    """The label for every parasite type selected on a capture, in stored order —
    each written into the Bemerkungen column since the template has no Parasit
    column of its own. Falls back to the raw code for any type not (yet) in the
    vocabulary, so a stray code never breaks the export."""
    return [_PARASIT_LABELS.get(code, code) for code in entry.parasites or []]


def _umstand(entry):
    """The row's Umstand: the Projekt's circumstance, displaced by ``08`` on a
    Tot-Fund — the one capture-level fact about the circumstance of *that* catch
    (ADR 0034). A Tot-Fund keeps its Bemerkung alongside the code."""
    if entry.is_dead_recovery:
        return TOT_FUND_UMSTAND
    return entry.project.circumstance if entry.project else None


def _zustand(entry):
    """The bird's condition when it left the Beringer's hand (ADR 0034): ``2`` on a
    Tot-Fund, ``8`` otherwise — and nothing at all for a *Ring vernichtet* row,
    where there was no bird whose condition could be asserted (the same reasoning
    that leaves that row's biometry empty)."""
    if entry.species and entry.species.special_kind == Species.SpecialKind.RING_DESTROYED:
        return None
    return TOT_FUND_ZUSTAND if entry.is_dead_recovery else DEFAULT_ZUSTAND


def _build_comment(entry):
    parts = [entry.comment] if entry.comment else []
    parts.extend(_parasite_labels(entry))
    if entry.has_hunger_stripes:
        parts.append("Hungerstreifen")
    if entry.has_brood_patch:
        parts.append("Brutfleck")
    if entry.has_cpl_plus:
        parts.append("CPL+")
    # Comma-joined, not space-joined (issue #406): several parasite labels are
    # themselves multi-word ("Rote Milben"), and recording more than one type is
    # routine — Zecken *and* Rote Milben is an ordinary finding. Joined by a bare
    # space, "Frischer Fang Rote Milben Zecke Brutfleck" gives the human reader at
    # the Meldestelle no token boundary, and this is the official Meldung.
    return ", ".join(parts) or None


# IWM header text → callable(entry) -> cell value (None = leave blank).
COLUMN_MAP = {
    # The ring's own issuing Zentrale — an EURING scheme code, never free prose.
    # The AUW backfill (ADR 0019) guarantees ``central`` is never null, so a
    # domestic ring still emits "AUW" while a foreign ring emits its own scheme
    # code (issue #230, US 20).
    "Ring": lambda e: e.ring.central.scheme_code,
    "Ringnummer": lambda e: f"{e.ring.size}{e.ring.number}",
    "Ringstatus": lambda e: e.bird_status.upper() if e.bird_status else None,
    "Art": lambda e: e.species.common_name_de,
    "Zusatzmarkierung": lambda e: ZUSATZMARKIERUNG,
    "Geschlecht": _sex_code,
    "Alter": lambda e: e.age_class,
    "Datum": lambda e: localtime(e.date_time).date(),
    "Uhrzeit": lambda e: localtime(e.date_time).time(),
    "Flügellänge": lambda e: e.wing_span,
    "Teilfederlänge": lambda e: e.feather_span,
    "Gewicht": lambda e: e.weight_gram,
    "Tarsus": lambda e: e.tarsus,
    "Fett": lambda e: _text_code(e.fat_deposit),
    "Muskel": lambda e: _text_code(e.muscle_class),
    "Intensität": lambda e: _text_code(e.small_feather_int),
    "Fortschritt": lambda e: e.small_feather_app,
    "Handschwingen": lambda e: _text_code(e.hand_wing),
    # Breeding indicators (issue #375): "J" or blank. The flags also stay as
    # text tokens in the Bemerkungen column (see ``_build_comment``).
    "Brutfleck": lambda e: _breeding_flag(e.has_brood_patch),
    "Kloake": lambda e: _breeding_flag(e.has_cpl_plus),
    "Netz": lambda e: _text_code(e.net_location),
    "Ort": lambda e: e.ringing_station.name if e.ringing_station else None,
    "Land": lambda e: e.ringing_station.country or None if e.ringing_station else None,
    "Region": lambda e: e.ringing_station.region or None if e.ringing_station else None,
    "Ortskodierung": (
        lambda e: e.ringing_station.place_code or None if e.ringing_station else None
    ),
    "Geo-Koordinaten": _geo_coordinates,
    "Umstand": _umstand,
    "Zustand": _zustand,
    "Fangmethode": lambda e: e.project.capture_method if e.project else None,
    "Lockmittel": lambda e: e.project.lure if e.project else None,
    "Bemerkungen": _build_comment,
    "BeringerIn": lambda e: e.staff.handle,
}


def build_iwm_workbook(entries) -> bytes:
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb[SHEET_NAME]

    header_to_col = {
        ws.cell(row=1, column=c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=c).value
    }
    missing = set(COLUMN_MAP) - set(header_to_col)
    if missing:
        raise RuntimeError(f"IWM template missing columns: {sorted(missing)}")

    # Clear all template data rows in the 32 real columns (deferred ones too),
    # so example rows — both their values and their leftover fills — don't leak
    # into the user's export.
    for r in range(2, ws.max_row + 1):
        for col in header_to_col.values():
            cell = ws.cell(row=r, column=col)
            cell.value = None
            cell.fill = _NO_FILL

    for row_idx, entry in enumerate(entries, start=2):
        # Nicht-Standard-Fang (ADR 0026): fill the whole row and blank the method
        # columns the Projekt supplies. A Tot-Fund is not filled — it carries its
        # own Umstand/Zustand codes instead (ADR 0034), which is also why its
        # Umstand is no longer Projekt-sourced and therefore survives the
        # combination of both Fangmarker.
        non_standard = entry.is_non_standard
        blank_columns = _project_sourced_columns(entry) if non_standard else frozenset()
        for header, getter in COLUMN_MAP.items():
            value = None if header in blank_columns else getter(entry)
            ws.cell(row=row_idx, column=header_to_col[header]).value = value
        if non_standard:
            for col in header_to_col.values():
                ws.cell(row=row_idx, column=col).fill = NON_STANDARD_FILL

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
