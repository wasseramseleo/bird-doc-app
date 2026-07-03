"""IWM import service (issue #120, PRD #113).

The other half of the IWM round-trip: where ``iwm_export.py`` turns a Projekt's
captures into an authentic ``Datenmeldung`` ``Fangdaten`` sheet, this turns such
a sheet back into captures in a chosen Projekt. It is the single ingestion code
path — exposed to Org-Admins through the ``import-iwm`` API action and (later)
reused by the demo-seed management command — and it creates captures through the
shared ``capture_service`` so imported rows obey exactly the same invariants as
form-entered ones (ADR 0006 org-scoped rings, ADR 0004 Sonderart rules).

Two phases on one upload (ADR 0013): a **dry-run** parses, validates and returns
an ``ImportPreview`` while writing nothing; a **commit** atomically creates the
importable captures, skips the blocking-error rows, and returns an
``ImportResult``. The report shapes are fixed by the PRD. Duplicate detection
populates ``duplicates``/``duplicatesSkipped`` (issue #122); the Projekt-method
reconciliation populates ``warnings`` (issue #124); the ``cap`` is enforced — an
upload whose data-row count exceeds ``ROW_CAP`` is rejected on both phases (issue
#125).

Unfamiliar Beringer and Stationen are auto-created rather than rejected (issue
#121). An unfamiliar Kürzel (resolved within the Organisation) becomes a
no-account Beringer (ADR 0001); an unfamiliar Ort (resolved by Ortskodierung /
name) becomes a Station built from the sheet's name / Ortskodierung / Region /
Land / coordinates. Every such creation is surfaced in the preview's
``toCreate`` for the Admin to approve, created only on commit (nothing on
dry-run), reported in ``ImportResult.{createdBeringer, createdStationen}``, and
attached to the captures that referenced it.

Beyond the spine columns (species, ring size+number, Beringer Kürzel, Station,
Datum+Uhrzeit, Ringstatus, Geschlecht, Alter) the importer ingests — added in
issue #123 — the Sonderarten and the authentic category codes. The Sonderart
names resolve to their Sonderart rows (*Ring Vernichtet* imports with all bird
data nulled by ``create_capture``, *Aves ignota* enforces its mandatory Bemerkung
as a blocking error via ``validate_capture``, ADR 0004), and the text category
codes (Fett/Muskel/Intensität/Handschwingen and the Netz number) are read into
their fields; ``Zusatzmarkierung="ZZ"`` (no additional marking) rides along
without a model field to land in. Ring numbers import as-is so historical Ring
identities are preserved (ADR 0006).

An unknown species, a missing required field (no ring number / no date) and a
blank Aves-ignota Bemerkung are blocking errors, as is a cross-Organisation
Kürzel collision (``_RowError``). Duplicate detection, the row cap and
Projekt-method warnings are enforced as described above.
"""

import re
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO

import openpyxl
from django.db import transaction
from django.utils.timezone import make_aware

from .capture_service import CaptureValidationError, create_capture, validate_capture
from .models import (
    AUW_SCHEME_CODE,
    Central,
    DataEntry,
    Ring,
    RingingStation,
    Scientist,
    Species,
    get_auw_central,
)
from .station_handle import derive_station_handle

SHEET_NAME = "Fangdaten"

# The authentic Fangdaten columns the importer needs to build a capture. A file
# missing any of them is structurally wrong and rejected before any row is read.
REQUIRED_HEADERS = frozenset(
    {
        "Ringnummer",
        "Ringstatus",
        "Art",
        "Geschlecht",
        "Alter",
        "Datum",
        "Uhrzeit",
        "Ort",
        "BeringerIn",
    }
)

# A single tunable row cap (a starting number — tune by measurement, ADR 0013).
# It bounds one synchronous upload: the preview reports it as ``cap`` and a file
# whose data-row count exceeds it is rejected on both preview and commit (issue
# #125) with guidance to split the file or bulk-load via the management command —
# never a silent partial import or truncation.
ROW_CAP = 5000

_VALID_RING_SIZES = frozenset(Ring.RingSizes.values)
_RINGNUMMER_RE = re.compile(r"^([A-Za-z]+)(\d+)$")

# Authentic IWM codes → model values (issue #120 AC: Geschlecht U/M/W, Ringstatus
# E/W understood). Alter is already an integer in the sheet.
_GESCHLECHT = {
    "U": DataEntry.Sex.UNKNOWN,
    "M": DataEntry.Sex.MALE,
    "W": DataEntry.Sex.FEMALE,
}
_RINGSTATUS = {
    "E": DataEntry.BirdStatus.FIRST_CATCH,
    "W": DataEntry.BirdStatus.RE_CATCH,
}


class IwmStructureError(Exception):
    """The upload is not a readable ``Datenmeldung`` — wrong file, no ``Fangdaten``
    sheet, or missing required headers. Raised before any row is processed so the
    caller can fast-fail with a clear message (HTTP 400)."""


class IwmRowCapExceeded(Exception):
    """The upload has more data rows than the import cap allows (ADR 0013, issue
    #125). A very large history must be split or bulk-loaded via the management
    command, never silently partial-imported or truncated. Raised before any row
    is resolved or written so both preview and commit reject it cleanly — the
    caller turns it into an HTTP 400 that signals the cap and points the Admin to
    the split / management-command path."""

    def __init__(self, count, limit):
        self.count = count
        self.limit = limit
        self.message = (
            f"Die Datei enthält {count} Datenzeilen und überschreitet die Obergrenze "
            f"von {limit} Zeilen pro Import. Bitte die Datei in kleinere Dateien "
            "aufteilen oder eine:n Operator:in bitten, den Bulk-Load per "
            "Management-Kommando auszuführen — es wurde nichts importiert und nichts "
            "abgeschnitten."
        )
        super().__init__(self.message)


def build_import_preview(content, project):
    """Dry-run: parse + validate ``content`` against ``project`` and return an
    ``ImportPreview``. Writes nothing — the unfamiliar Beringer and Stationen it
    would create are listed in ``toCreate`` but not yet created, and the
    Projekt-method adoption is deferred to commit."""
    rows, warnings, _adoptions, resolver = _analyze(content, project, create=False)
    seen = _existing_keys(project)
    importable = 0
    duplicates = 0
    errors = []
    for resolved in rows:
        if resolved.error is not None:
            errors.append({"row": resolved.row, "reason": resolved.error})
            continue
        key = _capture_key(resolved.kwargs)
        if key in seen:
            duplicates += 1
        else:
            seen.add(key)
            importable += 1
    return {
        "importable": importable,
        "duplicates": duplicates,
        "errors": errors,
        "warnings": warnings,
        "toCreate": {
            "beringer": resolver.created_beringer,
            "stationen": resolver.created_stationen,
        },
        "cap": {"limit": ROW_CAP, "exceeded": len(rows) > ROW_CAP},
    }


@transaction.atomic
def commit_import(content, project, *, enforce_cap=True):
    """Commit: create every importable capture atomically, skipping blocking-error
    rows and duplicates, and return an ``ImportResult``. All-or-nothing — an
    unexpected failure rolls the whole import back.

    A row whose capture key already exists in the Organisation (or was already
    imported earlier in this same file) is skipped as a duplicate and counted,
    never re-inserted — so re-importing a corrected file cannot double the data
    (issue #122). Unfamiliar Beringer and Stationen are created as part of the
    same transaction and reported in ``createdBeringer`` / ``createdStationen``
    (issue #121).

    ``enforce_cap`` is the one knob separating the two callers (ADR 0013). The
    ``import-iwm`` API leaves it ``True`` so an oversized synchronous upload is
    rejected with guidance (issue #125); the ``seed_demo_org`` management command
    sets it ``False`` to run the same parse → validate → create path for an
    ops-assisted over-cap backfill (a demo seed or a multi-year one-time history)
    without a background worker."""
    rows, _warnings, adoptions, resolver = _analyze(
        content, project, create=True, enforce_cap=enforce_cap
    )
    seen = _existing_keys(project)
    created = 0
    duplicates_skipped = 0
    errors = []
    for resolved in rows:
        if resolved.error is not None:
            errors.append({"row": resolved.row, "reason": resolved.error})
            continue
        key = _capture_key(resolved.kwargs)
        if key in seen:
            duplicates_skipped += 1
            continue
        try:
            create_capture(**resolved.kwargs)
            seen.add(key)
            created += 1
        except CaptureValidationError as exc:
            errors.append({"row": resolved.row, "reason": str(exc.message)})
    _adopt_context(project, adoptions)
    return {
        "created": created,
        "duplicatesSkipped": duplicates_skipped,
        "errors": errors,
        "createdBeringer": resolver.created_beringer,
        "createdStationen": resolver.created_stationen,
    }


def _capture_key(kwargs):
    """The duplicate-detection key of a resolved row: ring size + number and the
    exact capture datetime (issue #122). An Erstfang and its later Wiederfang
    share the ring but differ by datetime, so their keys differ and both import."""
    return (kwargs["ring_size"], kwargs["ring_number"], kwargs["date_time"])


def _existing_keys(project):
    """The set of capture keys already present in the Projekt's Organisation.

    Built once per import (pre-resolved lookup — ADR 0013) so row classification
    does not re-query. A resolved row whose key is in this set already exists and
    is skipped as a duplicate. Datetimes come back timezone-aware; an aware
    datetime hashes and compares by its UTC instant, so a row whose Datum+Uhrzeit
    the importer localises to the same wall clock matches its stored capture."""
    return {
        (size, number, date_time)
        for size, number, date_time in DataEntry.objects.filter(
            organization=project.organization
        ).values_list("ring__size", "ring__number", "date_time")
    }


class _ResolvedRow:
    """One data row resolved to either capture kwargs (``error is None``) or a
    blocking ``error`` reason, tagged with its 1-based sheet row number."""

    __slots__ = ("row", "kwargs", "error")

    def __init__(self, row, kwargs=None, error=None):
        self.row = row
        self.kwargs = kwargs
        self.error = error


class _RowError(Exception):
    """A per-row resolution failure raised from inside the resolver — an
    auto-create blocked by a constraint (e.g. a Kürzel already owned by another
    Organisation, ``Scientist.handle`` being globally unique). Caught in
    ``_resolve_row`` and turned into a blocking ``_ResolvedRow`` error so one bad
    row never aborts the whole atomic import."""

    def __init__(self, reason):
        self.reason = reason
        super().__init__(reason)


class _Resolver:
    """Org-scoped entity lookups, built once per import so row resolution does not
    re-query. Beringer and Stationen are small (org-scoped); species is looked up
    lazily and cached because the global table is huge.

    An unfamiliar Beringer (by Kürzel) or Station (by Ortskodierung / name) is
    *auto-created* rather than rejected (issue #121). ``create`` decides whether a
    resolution actually writes: a dry-run (``create=False``) only records what it
    *would* create in ``created_beringer`` / ``created_stationen``; a commit
    (``create=True``) creates the entity once and caches it so every referencing
    row shares it. Either way each unfamiliar entity is surfaced exactly once, in
    first-seen order.
    """

    def __init__(self, project, *, create):
        self.project = project
        self.org = project.organization
        self.create = create
        # The full EURING Zentralen list (global reference data, ~60 rows), loaded
        # once so the "Ring" column resolves without re-querying per row. AUW is
        # the fallback for an absent/blank cell (backward compatible).
        self.centrals = {c.scheme_code: c for c in Central.objects.all()}
        self.auw = self.centrals.get(AUW_SCHEME_CODE) or get_auw_central()
        self.beringer = {
            s.handle: s for s in Scientist.objects.filter(organization=self.org) if s.handle
        }
        stations = list(RingingStation.objects.filter(organization=self.org))
        self.station_by_code = {s.place_code: s for s in stations if s.place_code}
        self.station_by_name = {s.name: s for s in stations if s.name}
        self._species_cache = {}
        # Auto-creations surfaced in the preview / result (first-seen order,
        # deduplicated): the Kürzel of new Beringer and the names of new Stationen.
        self.created_beringer = []
        self.created_stationen = []
        # Per-import caches so a repeated unfamiliar reference resolves to the one
        # entity (a shared object on commit; ``None`` placeholder on dry-run).
        self._new_beringer = {}
        self._new_station = {}

    def species(self, common_name_de):
        if common_name_de not in self._species_cache:
            self._species_cache[common_name_de] = Species.objects.filter(
                common_name_de=common_name_de
            ).first()
        return self._species_cache[common_name_de]

    def central_for(self, code):
        """Resolve the ``Ring`` column's EURING scheme code to a ``Central``.

        An absent column or a blank cell means the domestic Austrian Vogelwarte
        (``AUW``) — old sheets and existing import workflows stay unbroken (US 25).
        A known seeded scheme code (matched case-insensitively — EURING codes are
        canonically upper-case) resolves to that Zentrale. An unknown code is
        returned as ``(None, message)`` so the row is rejected with a clear German
        message, surfacing a typo instead of corrupting data (US 24).

        Returns ``(central, None)`` on success or ``(None, error_message)``.
        """
        if not code:
            return self.auw, None
        central = self.centrals.get(code.strip().upper())
        if central is None:
            return None, _unknown_scheme_message(code)
        return central, None

    def beringer_for(self, kuerzel):
        """Resolve a Kürzel to its Beringer, auto-creating an unfamiliar one.

        A familiar Kürzel resolves to the existing org Beringer. An unfamiliar one
        is recorded in ``created_beringer`` (once) and, on commit, created as a
        no-account Beringer scoped to the Organisation (ADR 0001); on a dry-run
        nothing is written and ``None`` stands in as a placeholder.

        The familiarity map is org-scoped, but ``Scientist.handle`` is *globally*
        unique (models.py): a Kürzel already owned by a Beringer in another
        Organisation is unfamiliar here yet cannot be auto-created without
        violating that constraint. Rather than let the ``IntegrityError`` abort
        the whole atomic import (HTTP 500), such a row is a blocking error
        (``_RowError``) — the same for a dry-run preview and a commit."""
        existing = self.beringer.get(kuerzel)
        if existing is not None:
            return existing
        if kuerzel not in self._new_beringer:
            if Scientist.objects.filter(handle=kuerzel).exists():
                raise _RowError(
                    f"Beringer:in-Kürzel {kuerzel!r} ist bereits in einer anderen "
                    "Organisation vergeben und kann nicht angelegt werden."
                )
            self.created_beringer.append(kuerzel)
            self._new_beringer[kuerzel] = (
                Scientist.objects.create(handle=kuerzel, organization=self.org)
                if self.create
                else None
            )
        return self._new_beringer[kuerzel]

    def station_for(self, descriptors):
        """Resolve a Station by Ortskodierung / name, auto-creating an unfamiliar
        one from the sheet's descriptors.

        A Station is familiar when its Ortskodierung, or failing that its name,
        matches an existing org Station. An unfamiliar one is recorded in
        ``created_stationen`` (once, by name) and, on commit, created scoped to the
        Organisation from the sheet's name / Ortskodierung / Region / Land /
        coordinates; on a dry-run nothing is written."""
        place_code = descriptors["place_code"]
        name = descriptors["name"]
        if place_code and place_code in self.station_by_code:
            return self.station_by_code[place_code]
        if name and name in self.station_by_name:
            return self.station_by_name[name]
        key = place_code or name
        if key not in self._new_station:
            self.created_stationen.append(name or place_code)
            self._new_station[key] = self._create_station(descriptors) if self.create else None
        return self._new_station[key]

    def _create_station(self, descriptors):
        name = descriptors["name"] or descriptors["place_code"]
        handle = derive_station_handle(
            self.org, name, taken=lambda h: RingingStation.objects.filter(handle=h).exists()
        )
        station = RingingStation.objects.create(
            handle=handle,
            name=name,
            organization=self.org,
            place_code=descriptors["place_code"] or "",
            region=descriptors["region"] or "",
            country=descriptors["country"] or "",
            latitude=descriptors["latitude"],
            longitude=descriptors["longitude"],
        )
        # Later rows naming the same site resolve to this new Station too.
        if station.place_code:
            self.station_by_code[station.place_code] = station
        self.station_by_name[station.name] = station
        return station


def _analyze(content, project, *, create, enforce_cap=True):
    """Parse the workbook, resolve every data row, and reconcile the Projekt-scoped
    context columns.

    Returns ``(rows, warnings, adoptions, resolver)``: a ``_ResolvedRow`` per data
    row (never partial — a bad row becomes an error, not an exception), the
    non-blocking context warnings (Projekt-method mismatch / heterogeneous file),
    the ``{field: value}`` adoptions to write onto the Projekt on commit, and the
    ``_Resolver`` that carries the auto-created Beringer / Stationen. ``create``
    decides whether unfamiliar entities are actually written (commit) or only
    recorded for the preview (dry-run). Raises ``IwmStructureError`` for a
    structurally-wrong file and — when ``enforce_cap`` (the default, the API path)
    — ``IwmRowCapExceeded`` for an over-cap one, both before any row is resolved so
    nothing is written and both preview and commit reject cleanly (issue #125).
    The management command imports over-cap backfills with ``enforce_cap=False``
    (ADR 0013)."""
    header_index, data_rows = _read_fangdaten(content)
    if enforce_cap and len(data_rows) > ROW_CAP:
        raise IwmRowCapExceeded(len(data_rows), ROW_CAP)
    resolver = _Resolver(project, create=create)
    rows = [
        _resolve_row(values, header_index, row_num, project, resolver)
        for row_num, values in data_rows
    ]
    warnings, adoptions = _reconcile_context(data_rows, header_index, project)
    return rows, warnings, adoptions, resolver


# Fangmethode/Lockmittel/Umstand are Projekt properties (ADR 0002), never stored
# per capture. Each entry maps the file's Fangdaten column to the Projekt field it
# governs and the German label used in warnings.
_CONTEXT_COLUMNS = (
    ("Fangmethode", "capture_method", "Fangmethode"),
    ("Lockmittel", "lure", "Lockmittel"),
    ("Umstand", "circumstance", "Umstand"),
)


def _reconcile_context(data_rows, header_index, project):
    """Reconcile the file's Projekt-scoped context columns against the Projekt.

    The Projekt's value is authoritative and never overwritten when already set;
    the file's column is informational (ADR 0002). Per context column:

    * a **homogeneous** file value that differs from a **set** Projekt value → a
      non-blocking warning (the Projekt governs);
    * a homogeneous file value while the Projekt value is **unset** → adopt the
      file's value onto the Projekt (applied on commit);
    * a **heterogeneous** file (differing values across rows) → a warning, since
      the model cannot store a per-capture method.

    Returns ``(warnings, adoptions)``."""
    warnings = []
    adoptions = {}
    for column, field, label in _CONTEXT_COLUMNS:
        if column not in header_index:
            continue
        seen = [
            (row_num, value)
            for row_num, values in data_rows
            if (value := _clean(_cell(values, header_index, column))) is not None
        ]
        if not seen:
            continue  # the file says nothing about this column
        distinct = {value for _, value in seen}
        if len(distinct) > 1:
            first_value = seen[0][1]
            divergent_row = next(row_num for row_num, value in seen if value != first_value)
            warnings.append(
                {
                    "row": divergent_row,
                    "reason": (
                        f"{label}: uneinheitliche Werte in der Datei "
                        f"({', '.join(sorted(distinct))}). Die Methode ist eine "
                        "Projekt-Eigenschaft und kann nicht pro Fang gespeichert "
                        "werden; der Projektwert bleibt maßgeblich."
                    ),
                }
            )
            continue
        row_num, file_value = seen[0]
        project_value = (getattr(project, field) or "").strip()
        if not project_value:
            adoptions[field] = file_value
        elif project_value != file_value:
            warnings.append(
                {
                    "row": row_num,
                    "reason": (
                        f"{label}: Der Dateiwert „{file_value}“ weicht vom "
                        f"Projektwert „{project_value}“ ab; der Projektwert bleibt "
                        "maßgeblich."
                    ),
                }
            )
    return warnings, adoptions


def _adopt_context(project, adoptions):
    """Write the adopted context values onto the Projekt (commit only). No-op when
    there is nothing to adopt, so a plain import never touches the Projekt row."""
    if not adoptions:
        return
    for field, value in adoptions.items():
        setattr(project, field, value)
    project.save(update_fields=list(adoptions))


def _read_fangdaten(content):
    """Return ``(header_index, data_rows)`` for the ``Fangdaten`` sheet.

    ``header_index`` maps each header to its column offset; ``data_rows`` is a
    list of ``(sheet_row_number, values_tuple)`` for the non-empty data rows."""
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises a grab-bag of errors on junk input
        raise IwmStructureError(
            "Die Datei konnte nicht als Excel-Arbeitsmappe gelesen werden."
        ) from exc

    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise IwmStructureError(f"Das Blatt „{SHEET_NAME}“ fehlt in der Datei.")

    ws = wb[SHEET_NAME]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        wb.close()
        raise IwmStructureError(f"Das Blatt „{SHEET_NAME}“ ist leer.")

    header_index = {h: i for i, h in enumerate(header) if h}
    missing = REQUIRED_HEADERS - set(header_index)
    if missing:
        wb.close()
        raise IwmStructureError(
            "Der Datei fehlen erforderliche Spalten: " + ", ".join(sorted(missing)) + "."
        )

    data_rows = []
    for row_num, values in enumerate(rows_iter, start=2):
        if all(v is None for v in values):
            continue
        data_rows.append((row_num, values))
    wb.close()
    return header_index, data_rows


def _resolve_row(values, header_index, row_num, project, resolver):
    """Resolve one data row into capture kwargs or the first blocking error."""

    def text(header):
        return _clean(_cell(values, header_index, header))

    ringnummer = text("Ringnummer")
    if not ringnummer:
        return _ResolvedRow(row_num, error="Ringnummer fehlt.")
    # The "Ring" column names the ring's issuing Zentrale (an EURING scheme code);
    # absent/blank means the domestic AUW (backward compatible). An unknown code is
    # rejected so a typo surfaces instead of corrupting data (US 24).
    central, central_error = resolver.central_for(text("Ring"))
    if central_error is not None:
        return _ResolvedRow(row_num, error=central_error)
    parsed = _split_ringnummer(ringnummer, central)
    if parsed is None:
        # AUW keeps its historical message; a foreign row whose Ringnummer cannot
        # be split by the generic letters+digits regex (an exotic format) is
        # rejected with guidance to record it manually — never a silent
        # mis-import (US 23).
        if central.scheme_code == AUW_SCHEME_CODE:
            return _ResolvedRow(row_num, error=f"Ungültige Ringnummer: {ringnummer!r}.")
        return _ResolvedRow(
            row_num, error=_foreign_unsplittable_message(ringnummer, central.scheme_code)
        )
    ring_size, ring_number = parsed

    datum = _cell(values, header_index, "Datum")
    if _is_blank(datum):
        return _ResolvedRow(row_num, error="Datum fehlt.")
    date_time = _combine_datetime(datum, _cell(values, header_index, "Uhrzeit"))
    if date_time is None:
        return _ResolvedRow(row_num, error="Ungültiges Datum bzw. Uhrzeit.")

    art = text("Art")
    if not art:
        return _ResolvedRow(row_num, error="Art fehlt.")
    species = resolver.species(art)
    if species is None:
        return _ResolvedRow(row_num, error=f"Unbekannte Art (nicht in der Artenliste): {art!r}.")

    kuerzel = text("BeringerIn")
    if not kuerzel:
        return _ResolvedRow(row_num, error="Beringer:in fehlt.")

    place_code = text("Ortskodierung")
    name = text("Ort")
    if not place_code and not name:
        return _ResolvedRow(row_num, error="Ort bzw. Ortskodierung fehlt.")
    # An unfamiliar Kürzel / Ort is auto-created (issue #121); coordinates for a
    # new Station are parsed from the export's "lat, lon" Geo-Koordinaten. An
    # auto-create blocked by a constraint (e.g. a Kürzel already owned by another
    # Organisation) is a blocking row error, never a crash of the whole import.
    latitude, longitude = _parse_coordinates(text("Geo-Koordinaten"))
    try:
        staff = resolver.beringer_for(kuerzel)
        station = resolver.station_for(
            {
                "place_code": place_code,
                "name": name,
                "region": text("Region"),
                "country": text("Land"),
                "latitude": latitude,
                "longitude": longitude,
            }
        )
    except _RowError as exc:
        return _ResolvedRow(row_num, error=exc.reason)

    kwargs = {
        "species": species,
        "ring_size": ring_size,
        "ring_number": ring_number,
        # The ring's issuing Zentrale, resolved from the "Ring" column. The shared
        # capture service (#229) normalises the Größe against it (strict Austrian
        # under AUW, free text otherwise) and gates the status — no divergent
        # validation between the import and the capture write path (ADR 0019).
        "central": central,
        "staff": staff,
        "ringing_station": station,
        # Server-authoritative: the capture lands in the Projekt's Organisation,
        # never a client-chosen one (ADR 0005).
        "organization": project.organization,
        "project": project,
        "date_time": date_time,
        "comment": text("Bemerkungen"),
        "bird_status": _RINGSTATUS.get((text("Ringstatus") or "").upper()),
        "sex": _GESCHLECHT.get((text("Geschlecht") or "").upper()),
        "age_class": _parse_int(_cell(values, header_index, "Alter")),
        # Authentic IWM category codes carried as text in the sheet (issue #123).
        # Zusatzmarkierung="ZZ" means "no additional marking" — the model has no
        # field for it, so it is deliberately not read here.
        "fat_deposit": _parse_int(_cell(values, header_index, "Fett")),
        "muscle_class": _parse_int(_cell(values, header_index, "Muskel")),
        "small_feather_int": _parse_int(_cell(values, header_index, "Intensität")),
        "hand_wing": _parse_int(_cell(values, header_index, "Handschwingen")),
        "net_location": _parse_int(_cell(values, header_index, "Netz")),
        # Decimal biometrics + the moult Fortschritt the export emits (issue #176).
        # The columns are optional (never required); a blank or garbage cell parses
        # to None and never blocks the row, so export→import round-trips.
        "wing_span": _parse_decimal(_cell(values, header_index, "Flügellänge")),
        "feather_span": _parse_decimal(_cell(values, header_index, "Teilfederlänge")),
        "weight_gram": _parse_decimal(_cell(values, header_index, "Gewicht")),
        "tarsus": _parse_decimal(_cell(values, header_index, "Tarsus")),
        "small_feather_app": text("Fortschritt"),
    }
    # Run the shared creation invariants now (e.g. the *Aves ignota* mandatory
    # Bemerkung — ADR 0004) so the dry-run preview reports exactly the blocking
    # errors a commit would raise, instead of surfacing them only at commit time.
    try:
        validate_capture(kwargs["species"], kwargs["comment"])
    except CaptureValidationError as exc:
        return _ResolvedRow(row_num, error=str(exc.message))
    return _ResolvedRow(row_num, kwargs=kwargs)


def _cell(values, header_index, header):
    idx = header_index.get(header)
    if idx is None or idx >= len(values):
        return None
    return values[idx]


def _clean(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _is_blank(value):
    return value is None or (isinstance(value, str) and not value.strip())


def _unknown_scheme_message(code):
    """The rejection for an unknown ``Ring`` scheme code — a typo the Admin must
    fix, never a silent mis-import (US 24)."""
    return (
        f"Unbekannter Zentralen-Code {code!r} in der Spalte „Ring“. Bitte den Code "
        "prüfen oder den Eintrag manuell erfassen."
    )


def _foreign_unsplittable_message(ringnummer, scheme_code):
    """The rejection for a foreign Ringnummer the generic letters+digits regex
    cannot split — an exotic format the Admin must record by hand (US 23)."""
    return (
        f"Die Ringnummer {ringnummer!r} der Zentrale {scheme_code} konnte nicht "
        "automatisch in Ringgröße und Nummer aufgeteilt werden (ungewöhnliches "
        "Format). Bitte den Eintrag manuell erfassen."
    )


def _split_ringnummer(value, central):
    """Split a Ringnummer into ``(size, number)`` against its Zentrale, or ``None``.

    The size is the leading alphabetic run (matched greedily, so two-letter sizes
    like ``SA`` win over ``S``); the number is the trailing digits, kept as a
    string so leading zeros survive (ADR 0006). Under the Austrian Vogelwarte
    (``AUW``) the size must additionally be a known Austrian ring size — strict
    Austrian parsing, exactly as before the Zentrale slice. Under any other
    (foreign) Zentrale the generic letters+digits split alone governs: the leading
    letters become the free-text Größe (normalised against the Zentrale by the
    shared capture service) and the trailing digits the Nummer. Returns ``None``
    when the value does not match the generic letters+digits shape, or — for AUW —
    the size is not a known Austrian code."""
    match = _RINGNUMMER_RE.match(value)
    if not match:
        return None
    size = match.group(1).upper()
    if central.scheme_code == AUW_SCHEME_CODE and size not in _VALID_RING_SIZES:
        return None
    return size, match.group(2)


def _combine_datetime(datum, uhrzeit):
    """Combine the sheet's Datum + Uhrzeit into an aware datetime whose wall clock
    is what the ringer recorded. The naive value is localised to the project
    timezone (Europe/Vienna), the inverse of the export's ``localtime`` display —
    so a Datum/Uhrzeit round-trips to the same wall clock (issue #120 AC)."""
    date_part = _to_date(datum)
    if date_part is None:
        return None
    naive = datetime.combine(date_part, _to_time(uhrzeit) or time(0, 0))
    return make_aware(naive)


def _to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _to_time(value):
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    return None


def _parse_int(value):
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


# The two-place quantum matching the biometrics' ``DecimalField(decimal_places=2)``.
_TWO_PLACES = Decimal("0.01")


def _parse_decimal(value):
    """Parse a lenient biometric cell into a 2-place ``Decimal``, or ``None``.

    Accepts openpyxl numerics (``int``/``float``) and strings, treats both ``.``
    and ``,`` as the decimal separator (a German-locale Datenmeldung writes
    ``12,5``), and quantizes to two decimal places to match the model's
    ``DecimalField(decimal_places=2)``. Returns ``None`` on a blank or unparseable
    cell — lenient like ``_parse_int``, so one bad biometric cell never blocks a
    row. A ``float`` is routed through ``str`` first so its shortest decimal repr
    is preserved rather than its binary-float artefacts (``18.3`` → ``18.30``)."""
    if value is None:
        return None
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return Decimal(text).quantize(_TWO_PLACES)
    except (InvalidOperation, ValueError):
        return None


def _parse_coordinates(value):
    """Split the export's ``"lat, lon"`` Geo-Koordinaten (decimal degrees) into
    ``(latitude, longitude)`` as ``Decimal``. Returns ``(None, None)`` when the
    cell is blank or unparseable — coordinates are optional on an auto-created
    Station (issue #121)."""
    if not value:
        return None, None
    parts = value.split(",")
    if len(parts) != 2:
        return None, None
    try:
        return Decimal(parts[0].strip()), Decimal(parts[1].strip())
    except (InvalidOperation, ValueError):
        return None, None
