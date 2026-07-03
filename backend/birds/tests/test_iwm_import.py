"""IWM import — the end-to-end spine (issue #120, PRD #113).

Seam 1 (primary): the ``import-iwm`` API action. The tests drive it through the
DRF ``APIClient`` with in-test workbooks built the way ``test_iwm_export.py``
builds them, and assert external behaviour only — the HTTP response, the report
contents, and the captures that exist afterwards — never how parsing or
bulk-insert is wired internally.

Scope: dry-run preview + atomic commit on the same upload, Admin-only (403 for a
plain Mitglied, 404 for a foreign-tenant Projekt), structural fast-fail,
unknown-species / missing-field blocking errors, and wall-clock Datum/Uhrzeit
with Geschlecht ``U/M/W`` + integer Alter understood. The sibling slices are
covered too: the row cap (issue #125), auto-created Beringer/Stationen surfaced
in ``toCreate`` (issue #121), duplicate detection (issue #122), Projekt-method
warnings (issue #124), and — issue #123 — the Sonderarten (*Ring Vernichtet*
nulled, *Aves ignota* Bemerkung enforced) plus the authentic category codes.
"""

from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils.timezone import localtime, make_aware

from birds.iwm_export import build_iwm_workbook
from birds.iwm_import import _parse_decimal, commit_import
from birds.models import Central, DataEntry, Project, Ring, RingingStation, Scientist

PROJECTS_URL = "/api/birds/projects/"

# The committed authentic-format fixture (Illmitz/Neusiedlersee, 343 rows incl.
# Wiederfänge and the three Sonderart rows). Doubles as an import-feature fixture.
SAMPLE_IWM = Path(__file__).resolve().parent.parent / "demo" / "sample_iwm_illmitz.xlsx"

# The authentic IWM "Fangdaten" header row (a superset is fine; the importer
# reads the columns it needs by name), mirroring the export/sample layout. It
# carries the authentic category-code columns (Fett/Muskel/Intensität/
# Handschwingen/Netz) and Zusatzmarkierung so a test can drive full code fidelity.
HEADERS = [
    "Ring",
    "Ringnummer",
    "Ringstatus",
    "Zusatzmarkierung",
    "Art",
    "Geschlecht",
    "Alter",
    "Datum",
    "Uhrzeit",
    "Fett",
    "Muskel",
    "Intensität",
    "Handschwingen",
    "Netz",
    "Ortskodierung",
    "Ort",
    "Region",
    "Land",
    "Geo-Koordinaten",
    "Bemerkungen",
    "BeringerIn",
]


def _valid_row(species, scientist, ringing_station, **overrides):
    """A structurally-complete, resolvable Fangdaten row for tenant A."""
    row = {
        "Ring": "AUW",
        "Ringnummer": "V00604",
        "Ringstatus": "E",
        "Art": species.common_name_de,
        "Geschlecht": "U",
        "Alter": 3,
        "Datum": date(2026, 6, 30),
        "Uhrzeit": time(8, 15),
        "Ortskodierung": ringing_station.place_code or "",
        "Ort": ringing_station.name,
        "Region": "Burgenland",
        "Land": "Austria",
        "Bemerkungen": "",
        "BeringerIn": scientist.handle,
    }
    row.update(overrides)
    return row


def _workbook(rows, *, headers=HEADERS, sheet_name="Fangdaten"):
    """Serialise ``rows`` (list of {header: value}) into .xlsx bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    col = {h: i + 1 for i, h in enumerate(headers)}
    for r_i, row in enumerate(rows, start=2):
        for h, c in col.items():
            value = row.get(h)
            if value not in (None, ""):
                ws.cell(row=r_i, column=c, value=value)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(content, name="import.xlsx"):
    return SimpleUploadedFile(
        name,
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _import_url(project):
    return f"{PROJECTS_URL}{project.id}/import-iwm/"


@pytest.mark.django_db
def test_dry_run_returns_preview_and_writes_nothing(
    auth_client, scientist, ringing_station, project, species
):
    content = _workbook([_valid_row(species, scientist, ringing_station)])

    response = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    )

    assert response.status_code == 200, response.content
    body = response.json()
    assert body["importable"] == 1
    assert body["errors"] == []
    # A dry-run writes nothing — not a single capture.
    assert DataEntry.objects.count() == 0


@pytest.mark.django_db
def test_commit_imports_row_with_out_of_range_weight_despite_species_norm(
    auth_client, scientist, ringing_station, project, species, organization
):
    """The IWM import runs no plausibility check (PRD #245, ADR 0021): a row
    whose Gewicht is wildly outside the species' Artennorm band imports normally
    — historical rows may legitimately be "unusual" and must never warn or block."""
    from birds.models import SpeciesNorm

    SpeciesNorm.objects.create(
        species=species, organization=None, weight_mean=Decimal("9.1"), weight_sd=Decimal("0.82")
    )
    headers = [*HEADERS, "Gewicht"]
    content = _workbook(
        [_valid_row(species, scientist, ringing_station, **{"Gewicht": 250})],
        headers=headers,
    )

    response = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )

    assert response.status_code == 200, response.content
    body = response.json()
    assert body["created"] == 1
    assert body["errors"] == []
    assert DataEntry.objects.get().weight_gram == Decimal("250")


@pytest.mark.django_db
def test_commit_creates_importable_captures_scoped_to_project_org(
    auth_client, scientist, ringing_station, project, species, organization
):
    content = _workbook([_valid_row(species, scientist, ringing_station)])

    response = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )

    assert response.status_code == 200, response.content
    body = response.json()
    assert body["created"] == 1
    assert body["errors"] == []

    entry = DataEntry.objects.get()
    assert entry.species == species
    assert entry.staff == scientist
    assert entry.ringing_station == ringing_station
    assert entry.project == project
    assert entry.ring.size == "V"
    assert entry.ring.number == "00604"
    # Server-authoritative: scoped to the Projekt's Organisation, and its Ring is
    # created within that Organisation (ADR 0005, 0006).
    assert entry.organization == organization
    assert entry.ring.organization == organization


@pytest.mark.django_db
def test_commit_understands_geschlecht_letters_ringstatus_and_integer_alter(
    auth_client, scientist, ringing_station, project, species
):
    content = _workbook(
        [
            _valid_row(
                species,
                scientist,
                ringing_station,
                Ringnummer="V00701",
                Geschlecht="W",
                Alter=5,
                Ringstatus="W",
            )
        ]
    )

    response = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )

    assert response.status_code == 200, response.content
    entry = DataEntry.objects.get()
    assert entry.sex == DataEntry.Sex.FEMALE
    assert entry.age_class == 5
    assert entry.bird_status == DataEntry.BirdStatus.RE_CATCH


@pytest.mark.django_db
def test_datum_uhrzeit_import_as_wall_clock(
    auth_client, scientist, ringing_station, project, species
):
    # The sheet records a Vienna wall clock; the stored capture must display back
    # the same Datum/Uhrzeit under the export's localtime (issue #120 AC).
    content = _workbook(
        [
            _valid_row(
                species,
                scientist,
                ringing_station,
                Datum=date(2026, 7, 1),
                Uhrzeit=time(1, 0),
            )
        ]
    )

    response = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )

    assert response.status_code == 200, response.content
    entry = DataEntry.objects.get()
    local = localtime(entry.date_time)
    assert local.date() == date(2026, 7, 1)
    assert local.time() == time(1, 0)


# --- Admin-only (mirrors export-iwm): 403 Mitglied, 404 foreign tenant --------


@pytest.mark.django_db
def test_plain_mitglied_cannot_import(
    mitglied_client, mitglied_scientist, project, scientist, ringing_station, species
):
    # Mara is a member of tenant A but a plain Mitglied: the import is a
    # privileged bulk write, refused with a clear message (never a bare 403).
    project.scientists.add(mitglied_scientist)
    content = _workbook([_valid_row(species, scientist, ringing_station)])

    response = mitglied_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    )

    assert response.status_code == 403
    assert "Administrator" in response.json().get("detail", "")
    assert DataEntry.objects.count() == 0


@pytest.mark.django_db
def test_foreign_tenant_project_is_404(
    auth_client_b, scientist_b, project, scientist, ringing_station, species
):
    # Bruno is an Admin of tenant B; tenant A's Projekt is invisible to him, so
    # importing into it is a 404 (the row is absent), not a 403.
    content = _workbook([_valid_row(species, scientist, ringing_station)])

    response = auth_client_b.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )

    assert response.status_code == 404
    assert DataEntry.objects.count() == 0


# --- Structural fast-fail -----------------------------------------------------


@pytest.mark.django_db
def test_missing_fangdaten_sheet_fails_fast(auth_client, scientist, project):
    content = _workbook([{}], sheet_name="Tabelle1")

    response = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    )

    assert response.status_code == 400
    assert "Fangdaten" in str(response.json())


@pytest.mark.django_db
def test_missing_required_header_fails_fast(
    auth_client, scientist, ringing_station, project, species
):
    headers = [h for h in HEADERS if h != "Ringnummer"]
    content = _workbook([_valid_row(species, scientist, ringing_station)], headers=headers)

    response = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    )

    assert response.status_code == 400
    assert "Ringnummer" in str(response.json())


# --- Per-row blocking errors --------------------------------------------------


@pytest.mark.django_db
def test_unknown_species_is_a_blocking_error(
    auth_client, scientist, ringing_station, project, species
):
    rows = [
        _valid_row(species, scientist, ringing_station),  # row 2 — importable
        _valid_row(
            species,
            scientist,
            ringing_station,
            Ringnummer="V00777",
            Art="Ferkelvogel Nichtexistent",
        ),  # row 3 — unknown species
    ]
    content = _workbook(rows)

    preview = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    ).json()

    assert preview["importable"] == 1
    assert [e["row"] for e in preview["errors"]] == [3]
    assert "Ferkelvogel" in preview["errors"][0]["reason"]
    assert DataEntry.objects.count() == 0

    # On commit the bad row is skipped, the good one imported.
    result = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    ).json()
    assert result["created"] == 1
    assert [e["row"] for e in result["errors"]] == [3]
    assert DataEntry.objects.count() == 1


@pytest.mark.django_db
def test_missing_ring_number_and_missing_date_are_blocking_errors(
    auth_client, scientist, ringing_station, project, species
):
    rows = [
        _valid_row(species, scientist, ringing_station, Ringnummer=""),  # row 2
        _valid_row(species, scientist, ringing_station, Ringnummer="V00888", Datum=""),  # row 3
    ]
    content = _workbook(rows)

    preview = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    ).json()

    assert preview["importable"] == 0
    reasons = {e["row"]: e["reason"] for e in preview["errors"]}
    assert set(reasons) == {2, 3}
    assert "Ringnummer" in reasons[2]
    assert "Datum" in reasons[3]


# --- Row cap: oversized files rejected with guidance (issue #125) -------------


@pytest.mark.django_db
def test_over_cap_file_is_rejected_at_preview_with_cap_signalled(
    monkeypatch, auth_client, scientist, ringing_station, project, species
):
    # A very large history must be split or bulk-loaded, never silently
    # partial-imported or truncated (ADR 0013). With the cap lowered to 2 a
    # 3-row file is over the cap: preview rejects it, signals the cap and writes
    # nothing.
    monkeypatch.setattr("birds.iwm_import.ROW_CAP", 2)
    rows = [
        _valid_row(species, scientist, ringing_station, Ringnummer=f"V{600 + i:05d}")
        for i in range(3)
    ]
    content = _workbook(rows)

    response = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    )

    assert response.status_code == 400, response.content
    body = response.json()
    assert body["cap"] == {"limit": 2, "exceeded": True}
    # The message points the Admin to the split / management-command path.
    message = str(body["file"])
    assert "aufteilen" in message
    assert "Management-Kommando" in message
    assert DataEntry.objects.count() == 0


@pytest.mark.django_db
def test_over_cap_file_is_rejected_on_commit_and_writes_nothing(
    monkeypatch, auth_client, scientist, ringing_station, project, species
):
    # Skipping the preview and committing an over-cap file straight away must
    # also be refused — the cap is enforced on both phases, nothing is written.
    monkeypatch.setattr("birds.iwm_import.ROW_CAP", 2)
    rows = [
        _valid_row(species, scientist, ringing_station, Ringnummer=f"V{700 + i:05d}")
        for i in range(3)
    ]
    content = _workbook(rows)

    response = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )

    assert response.status_code == 400, response.content
    assert response.json()["cap"] == {"limit": 2, "exceeded": True}
    assert DataEntry.objects.count() == 0


@pytest.mark.django_db
def test_file_at_the_cap_previews_and_commits_normally(
    monkeypatch, auth_client, scientist, ringing_station, project, species
):
    # A file exactly at the cap is fine: it previews with the cap reported but
    # not exceeded, and commits every row.
    monkeypatch.setattr("birds.iwm_import.ROW_CAP", 2)
    rows = [
        _valid_row(species, scientist, ringing_station, Ringnummer=f"V{800 + i:05d}")
        for i in range(2)
    ]
    content = _workbook(rows)

    preview = auth_client.post(_import_url(project), {"file": _upload(content)}, format="multipart")
    assert preview.status_code == 200, preview.content
    assert preview.json()["cap"] == {"limit": 2, "exceeded": False}
    assert preview.json()["importable"] == 2
    assert DataEntry.objects.count() == 0

    result = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )
    assert result.status_code == 200, result.content
    assert result.json()["created"] == 2
    assert DataEntry.objects.count() == 2


# --- Auto-create unfamiliar Beringer & Stationen, surfaced in the preview -----
# (issue #121). Replaces the spine's temporary "unknown Beringer/Station is an
# error" behaviour: an unfamiliar Kürzel becomes a no-account Beringer (ADR 0001)
# and an unfamiliar Ort a new Station, each surfaced in ``toCreate`` for the
# Admin to approve, created only on commit, and attached to the referencing rows.


@pytest.mark.django_db
def test_dry_run_lists_unfamiliar_beringer_and_station_in_to_create(
    auth_client, scientist, ringing_station, project, species
):
    content = _workbook(
        [
            _valid_row(
                species,
                scientist,
                ringing_station,
                BeringerIn="XZY",
                Ort="Feldstation Nord",
                Ortskodierung="FN01",
                Region="Niederösterreich",
                Land="Austria",
            ),
        ]
    )

    preview = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    ).json()

    # The row is importable — the unfamiliar Kürzel and Ort are surfaced for the
    # Admin to approve, not rejected as errors.
    assert preview["importable"] == 1
    assert preview["errors"] == []
    assert preview["toCreate"]["beringer"] == ["XZY"]
    assert preview["toCreate"]["stationen"] == ["Feldstation Nord"]
    # A dry-run writes nothing — not the capture, and not the entities either.
    assert DataEntry.objects.count() == 0
    assert not Scientist.objects.filter(handle="XZY").exists()
    assert not RingingStation.objects.filter(place_code="FN01").exists()


@pytest.mark.django_db
def test_commit_auto_creates_beringer_and_station_and_attaches_them(
    auth_client, scientist, ringing_station, project, species, organization
):
    content = _workbook(
        [
            _valid_row(
                species,
                scientist,
                ringing_station,
                BeringerIn="XZY",
                Ort="Feldstation Nord",
                Ortskodierung="FN01",
                Region="Niederösterreich",
                Land="Austria",
            ),
        ]
    )

    result = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    ).json()

    assert result["created"] == 1
    assert result["createdBeringer"] == ["XZY"]
    assert result["createdStationen"] == ["Feldstation Nord"]

    # The unfamiliar Kürzel is now a no-account Beringer scoped to the Org (ADR 0001).
    beringer = Scientist.objects.get(handle="XZY")
    assert beringer.organization == organization
    assert beringer.user is None

    # The unfamiliar Ort is now a Station scoped to the Org, built from the sheet.
    station = RingingStation.objects.get(place_code="FN01")
    assert station.organization == organization
    assert station.name == "Feldstation Nord"
    assert station.region == "Niederösterreich"
    assert station.country == "Austria"

    # …and the imported capture is attached to both auto-created entities.
    entry = DataEntry.objects.get()
    assert entry.staff == beringer
    assert entry.ringing_station == station


@pytest.mark.django_db
def test_repeated_unfamiliar_beringer_is_created_once_and_shared(
    auth_client, scientist, ringing_station, project, species
):
    # Two rows naming the same unfamiliar Kürzel: it is surfaced once, created
    # once, and both captures share the one Beringer (no duplicate junk records).
    rows = [
        _valid_row(species, scientist, ringing_station, Ringnummer="V00601", BeringerIn="QQ"),
        _valid_row(species, scientist, ringing_station, Ringnummer="V00602", BeringerIn="QQ"),
    ]
    content = _workbook(rows)

    preview = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    ).json()
    assert preview["toCreate"]["beringer"] == ["QQ"]

    result = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    ).json()
    assert result["created"] == 2
    assert result["createdBeringer"] == ["QQ"]

    assert Scientist.objects.filter(handle="QQ").count() == 1
    beringer = Scientist.objects.get(handle="QQ")
    assert DataEntry.objects.filter(staff=beringer).count() == 2


@pytest.mark.django_db
def test_commit_auto_created_station_parses_geo_koordinaten(
    auth_client, scientist, ringing_station, project, species, organization
):
    # An unfamiliar Ort carrying a "lat, lon" Geo-Koordinaten cell: the parsed
    # decimal coordinates land on the auto-created Station (issue #121 AC).
    content = _workbook(
        [
            _valid_row(
                species,
                scientist,
                ringing_station,
                Ort="Feldstation Süd",
                Ortskodierung="FS02",
                Region="Steiermark",
                Land="Austria",
                **{"Geo-Koordinaten": "47.123456, 15.654321"},
            ),
        ]
    )

    result = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    ).json()

    assert result["created"] == 1
    assert result["createdStationen"] == ["Feldstation Süd"]

    station = RingingStation.objects.get(place_code="FS02")
    assert station.latitude == Decimal("47.123456")
    assert station.longitude == Decimal("15.654321")

    # …and the capture is attached to the coordinate-carrying Station.
    entry = DataEntry.objects.get()
    assert entry.ringing_station == station


@pytest.mark.django_db
def test_cross_org_kuerzel_collision_is_a_row_error_not_a_crash(
    auth_client, scientist, ringing_station, project, species, scientist_b
):
    # Bruno's Kürzel "BRU" is owned by tenant B. It is unfamiliar to tenant A
    # (the familiarity map is org-scoped) but Scientist.handle is globally unique,
    # so auto-creating it would violate the constraint. Instead of crashing the
    # whole atomic import (HTTP 500, rollback), the row is a blocking error — the
    # same on the dry-run preview and on commit.
    content = _workbook(
        [_valid_row(species, scientist, ringing_station, BeringerIn=scientist_b.handle)]
    )

    preview = auth_client.post(_import_url(project), {"file": _upload(content)}, format="multipart")
    assert preview.status_code == 200, preview.content
    preview_body = preview.json()
    assert preview_body["importable"] == 0
    assert [e["row"] for e in preview_body["errors"]] == [2]
    assert scientist_b.handle in preview_body["errors"][0]["reason"]
    # The colliding Kürzel is not offered for creation.
    assert preview_body["toCreate"]["beringer"] == []

    result = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )
    assert result.status_code == 200, result.content
    result_body = result.json()
    assert result_body["created"] == 0
    assert [e["row"] for e in result_body["errors"]] == [2]
    # No capture created, and tenant B's Beringer is untouched.
    assert DataEntry.objects.count() == 0
    assert Scientist.objects.filter(handle=scientist_b.handle).count() == 1


# --- Duplicate detection & recapture (issue #122) -----------------------------
# A row is a duplicate when a capture already exists in the Organisation with the
# same capture key = (ring size + number, date + time). Duplicates are skipped
# (never re-inserted) and counted; an Erstfang and its later Wiederfang share a
# ring but differ by datetime, so both import; re-importing an already-loaded
# file recognises every row as a duplicate, making fix-a-few-and-re-import safe.


def _seed_capture(
    species,
    scientist,
    ringing_station,
    project,
    organization,
    *,
    size="V",
    number="00604",
    when=datetime(2026, 6, 30, 8, 15),
):
    """An existing capture in tenant A whose capture key (ring size+number,
    Datum+Uhrzeit) a matching import row must recognise as a duplicate. Its
    datetime is stored the way the importer combines Datum+Uhrzeit — the sheet's
    Vienna wall clock — so the round-trip key matches exactly."""
    ring = Ring.objects.create(size=size, number=number, organization=organization)
    return DataEntry.objects.create(
        species=species,
        ring=ring,
        staff=scientist,
        ringing_station=ringing_station,
        project=project,
        organization=organization,
        date_time=make_aware(when),
    )


@pytest.mark.django_db
def test_duplicate_row_by_capture_key_is_reported_and_writes_nothing(
    auth_client, scientist, ringing_station, project, species, organization
):
    # A capture with this exact ring (V00604) and datetime already exists.
    _seed_capture(species, scientist, ringing_station, project, organization)
    content = _workbook([_valid_row(species, scientist, ringing_station)])

    body = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    ).json()

    assert body["importable"] == 0
    assert body["duplicates"] == 1
    assert body["errors"] == []
    # The dry-run wrote nothing — still just the one pre-existing capture.
    assert DataEntry.objects.count() == 1


@pytest.mark.django_db
def test_commit_skips_duplicate_and_reports_duplicates_skipped(
    auth_client, scientist, ringing_station, project, species, organization
):
    _seed_capture(species, scientist, ringing_station, project, organization)
    content = _workbook([_valid_row(species, scientist, ringing_station)])

    body = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    ).json()

    assert body["created"] == 0
    assert body["duplicatesSkipped"] == 1
    assert body["errors"] == []
    # The duplicate was skipped, not re-inserted — no doubling.
    assert DataEntry.objects.count() == 1


@pytest.mark.django_db
def test_erstfang_and_later_wiederfang_both_import(
    auth_client, scientist, ringing_station, project, species, organization
):
    # Same ring (V00604), different datetime: the Erstfang and its later
    # Wiederfang have distinct capture keys, so neither is a duplicate of the
    # other and both import — recapture history preserved.
    rows = [
        _valid_row(
            species,
            scientist,
            ringing_station,
            Ringstatus="E",
            Datum=date(2026, 6, 30),
            Uhrzeit=time(8, 15),
        ),
        _valid_row(
            species,
            scientist,
            ringing_station,
            Ringstatus="W",
            Datum=date(2026, 7, 15),
            Uhrzeit=time(9, 0),
        ),
    ]
    content = _workbook(rows)

    body = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    ).json()

    assert body["created"] == 2
    assert body["duplicatesSkipped"] == 0
    assert DataEntry.objects.count() == 2
    ring = Ring.objects.get(size="V", number="00604", organization=organization)
    assert DataEntry.objects.filter(ring=ring).count() == 2


@pytest.mark.django_db
def test_reimporting_the_same_file_skips_every_row_as_duplicate(
    auth_client, scientist, ringing_station, project, species
):
    # Two distinct captures (different rings). Fix-a-few-and-re-import safety:
    # the second run must recognise every row as already loaded and add nothing.
    rows = [
        _valid_row(species, scientist, ringing_station, Ringnummer="V00604"),
        _valid_row(species, scientist, ringing_station, Ringnummer="V00701"),
    ]
    content = _workbook(rows)

    first = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    ).json()
    assert first["created"] == 2
    assert DataEntry.objects.count() == 2

    preview = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    ).json()
    assert preview["importable"] == 0
    assert preview["duplicates"] == 2

    result = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    ).json()
    assert result["created"] == 0
    assert result["duplicatesSkipped"] == 2
    # No doubling — still exactly the two original captures.
    assert DataEntry.objects.count() == 2


@pytest.mark.django_db
def test_two_identical_rows_in_one_file_import_once(
    auth_client, scientist, ringing_station, project, species
):
    # Two rows with the same capture key within a single upload: the first
    # imports, the second is recognised as a duplicate of it — one file never
    # doubles its own data.
    row = _valid_row(species, scientist, ringing_station)
    content = _workbook([dict(row), dict(row)])

    body = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    ).json()

    assert body["created"] == 1
    assert body["duplicatesSkipped"] == 1
    assert DataEntry.objects.count() == 1


# --- Projekt-method precedence & non-blocking warnings (issue #124) ------------
# Fangmethode/Lockmittel/Umstand are Projekt properties (ADR 0002), never stored
# per capture. The selected Projekt's values govern; the file's columns are
# informational and drive the warnings channel, which is distinct from the
# blocking errors channel and never stops the import.


@pytest.mark.django_db
def test_homogeneous_file_disagreeing_with_set_method_warns_but_imports(
    auth_client, scientist, ringing_station, project, species
):
    # The Projekt's Fangmethode is set (default "M" — Japannetz). A homogeneous
    # file whose Fangmethode says "H" disagrees: a non-blocking warning is raised,
    # the Projekt value governs, and the import still proceeds.
    assert project.capture_method == Project.CaptureMethod.MIST_NET
    content = _workbook(
        [_valid_row(species, scientist, ringing_station, Fangmethode="H")],
        headers=[*HEADERS, "Fangmethode"],
    )

    preview = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    ).json()

    assert preview["importable"] == 1
    assert preview["errors"] == []  # the mismatch is a warning, never an error
    assert len(preview["warnings"]) == 1
    warning = preview["warnings"][0]
    assert warning["row"] == 2
    assert "Fangmethode" in warning["reason"]

    # Non-blocking: commit imports the row and leaves the Projekt value untouched.
    result = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    ).json()
    assert result["created"] == 1
    project.refresh_from_db()
    assert project.capture_method == Project.CaptureMethod.MIST_NET


@pytest.mark.django_db
def test_unset_projekt_method_is_adopted_from_homogeneous_file(
    auth_client, scientist, ringing_station, project, species
):
    # The Projekt's Lockmittel is unset; a homogeneous file supplies "A" (Futter).
    # There is no conflict — so no warning — and the file's value is adopted onto
    # the Projekt. Adoption is a write, so it happens only on commit.
    project.lure = ""
    project.save(update_fields=["lure"])
    content = _workbook(
        [
            _valid_row(species, scientist, ringing_station, Lockmittel="A"),
            _valid_row(species, scientist, ringing_station, Ringnummer="V00702", Lockmittel="A"),
        ],
        headers=[*HEADERS, "Lockmittel"],
    )

    preview = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    ).json()
    assert preview["warnings"] == []  # adoption is not a conflict
    project.refresh_from_db()
    assert project.lure == ""  # dry-run adopts nothing

    result = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    ).json()
    assert result["created"] == 2
    project.refresh_from_db()
    assert project.lure == "A"


@pytest.mark.django_db
def test_heterogeneous_file_warns_and_stores_no_per_row_method(
    auth_client, scientist, ringing_station, project, species
):
    # The file's Fangmethode differs across rows ("H" then "M"). The model cannot
    # store a per-capture method, so this is a non-blocking warning; both rows
    # still import. The Projekt is left unset to prove a heterogeneous file adopts
    # nothing (it warns instead of guessing which value governs).
    project.capture_method = ""
    project.save(update_fields=["capture_method"])
    content = _workbook(
        [
            _valid_row(species, scientist, ringing_station, Fangmethode="H"),
            _valid_row(species, scientist, ringing_station, Ringnummer="V00703", Fangmethode="M"),
        ],
        headers=[*HEADERS, "Fangmethode"],
    )

    preview = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    ).json()
    assert preview["importable"] == 2
    assert preview["errors"] == []
    assert len(preview["warnings"]) == 1
    warning = preview["warnings"][0]
    assert warning["row"] == 3  # the row where the value diverges
    assert "Fangmethode" in warning["reason"]

    result = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    ).json()
    assert result["created"] == 2
    project.refresh_from_db()
    assert project.capture_method == ""  # heterogeneous file adopts nothing


# --- Authentic IWM codes beyond the spine (issue #123) ------------------------


@pytest.mark.django_db
def test_commit_ingests_authentic_category_codes(
    auth_client, scientist, ringing_station, project, species
):
    # The authentic sheet carries the category fields as text codes (Fett, Muskel,
    # Intensität, Handschwingen) and the net number (Netz); Zusatzmarkierung="ZZ"
    # (no additional marking) rides along. Each lands in its model field.
    content = _workbook(
        [
            _valid_row(
                species,
                scientist,
                ringing_station,
                Zusatzmarkierung="ZZ",
                Fett="4",
                Muskel="2",
                Intensität="1",
                Handschwingen="3",
                Netz="7",
            )
        ]
    )

    response = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )

    assert response.status_code == 200, response.content
    assert response.json()["created"] == 1
    entry = DataEntry.objects.get()
    assert entry.fat_deposit == 4
    assert entry.muscle_class == 2
    assert entry.small_feather_int == 1
    assert entry.hand_wing == 3
    assert entry.net_location == 7


# --- Sonderarten (issue #123) -------------------------------------------------


@pytest.mark.django_db
def test_ring_vernichtet_imports_with_bird_data_nulled_and_ring_preserved(
    auth_client, scientist, ringing_station, project, sentinel_species
):
    # A destroyed-ring row still consumes its Ring identity but carries no bird
    # data: whatever codes ride along on the row are forced null server-side
    # (ADR 0004), and its Ringnummer is preserved as-is.
    content = _workbook(
        [
            _valid_row(
                sentinel_species,
                scientist,
                ringing_station,
                Ringnummer="V04211",
                Ringstatus="",
                Geschlecht="",
                Alter="",
                Fett="3",
                Muskel="2",
                Netz="4",
                Bemerkungen="Ring beim Anlegen deformiert, vernichtet",
            )
        ]
    )

    response = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )

    assert response.status_code == 200, response.content
    assert response.json()["created"] == 1
    entry = DataEntry.objects.get()
    assert entry.species == sentinel_species
    # Ring identity preserved as-is (ADR 0006) — no renumbering.
    assert entry.ring.size == "V"
    assert entry.ring.number == "04211"
    # Every bird-data field nulled, whatever the sheet carried.
    assert entry.sex is None
    assert entry.age_class is None
    assert entry.bird_status is None
    assert entry.fat_deposit is None
    assert entry.muscle_class is None
    assert entry.net_location is None


@pytest.mark.django_db
def test_aves_ignota_requires_bemerkung_blank_is_blocking_error(
    auth_client, scientist, ringing_station, project, aves_ignota_species
):
    # An unlisted rarity (Aves ignota) must always be described: a row with a
    # Bemerkung imports; one with a blank Bemerkung is a blocking error — visible
    # already in the dry-run preview and skipped on commit (ADR 0004).
    rows = [
        _valid_row(
            aves_ignota_species,
            scientist,
            ringing_station,
            Ringnummer="V05001",
            Bemerkungen="Unbestimmter Acrocephalus, Fotos an Vogelwarte",
        ),  # row 2 — importable
        _valid_row(
            aves_ignota_species,
            scientist,
            ringing_station,
            Ringnummer="V05002",
            Bemerkungen="",
        ),  # row 3 — blank Bemerkung, blocking
    ]
    content = _workbook(rows)

    preview = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    ).json()
    assert preview["importable"] == 1
    assert [e["row"] for e in preview["errors"]] == [3]
    assert "Bemerkung" in preview["errors"][0]["reason"]
    assert DataEntry.objects.count() == 0

    result = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    ).json()
    assert result["created"] == 1
    assert [e["row"] for e in result["errors"]] == [3]
    entry = DataEntry.objects.get()
    assert entry.species == aves_ignota_species
    assert entry.comment == "Unbestimmter Acrocephalus, Fotos an Vogelwarte"


@pytest.mark.django_db
def test_sonderart_names_resolve_to_sonderart_rows_not_unknown_species(
    auth_client, scientist, ringing_station, project, sentinel_species, aves_ignota_species
):
    # The Sonderart names are real Species rows, not typos: they resolve to their
    # Sonderart rows and must never be flagged "unknown species (not in the
    # Artenliste)" the way a genuine unlisted name (row 4) is.
    rows = [
        _valid_row(
            sentinel_species,
            scientist,
            ringing_station,
            Ringnummer="V06001",
            Ringstatus="",
            Bemerkungen="vernichtet",
        ),  # row 2 — Ring Vernichtet
        _valid_row(
            aves_ignota_species,
            scientist,
            ringing_station,
            Ringnummer="V06002",
            Bemerkungen="Bestimmung unsicher",
        ),  # row 3 — Aves ignota
        _valid_row(
            species=type("S", (), {"common_name_de": "Ferkelvogel Nichtexistent"})(),
            scientist=scientist,
            ringing_station=ringing_station,
            Ringnummer="V06003",
        ),  # row 4 — genuinely unknown
    ]
    content = _workbook(rows)

    preview = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    ).json()

    assert preview["importable"] == 2
    assert [e["row"] for e in preview["errors"]] == [4]
    assert "Unbekannte Art" in preview["errors"][0]["reason"]


# --- The committed authentic-IWM fixture (issue #123) -------------------------


def _sonderart_rows_from_fixture():
    """Read the fixture's three Sonderart rows keyed by Ringnummer, so the test
    asserts against the real committed sheet rather than hardcoded values."""
    wb = openpyxl.load_workbook(SAMPLE_IWM, data_only=True)
    ws = wb["Fangdaten"]
    headers = [c.value for c in ws[1]]
    sonderart = {}
    for r in ws.iter_rows(min_row=2):
        row = dict(zip(headers, [c.value for c in r], strict=False))
        if row.get("Art") in ("Ring Vernichtet", "Art nicht in der Liste (Aves ignota)"):
            sonderart[row["Ringnummer"]] = row
    wb.close()
    return sonderart


def _capture_for_ringnummer(ringnummer):
    """The imported capture whose Ring identity is ``ringnummer`` (e.g. ``V18245``)."""
    size = "".join(ch for ch in ringnummer if ch.isalpha())
    number = ringnummer[len(size) :]
    return DataEntry.objects.get(ring__size=size, ring__number=number)


@pytest.mark.django_db
def test_sample_fixture_imports_sonderart_rows_and_authentic_codes(
    auth_client, scientist, project, organization
):
    # The Sonderart rows reference Beringer JGR/MWA/SBA at Stationen NS01/NS02/NS03;
    # pre-seed those so the three rows resolve (auto-create is a later slice).
    for kuerzel in ("JGR", "MWA", "SBA"):
        Scientist.objects.create(handle=kuerzel, organization=organization)
    for code in ("NS01", "NS02", "NS03"):
        RingingStation.objects.create(
            handle=f"ST-{code}", name=f"Illmitz {code}", place_code=code, organization=organization
        )
    fixture = _sonderart_rows_from_fixture()
    assert len(fixture) == 3  # one Ring Vernichtet + two Aves ignota

    upload = _upload(SAMPLE_IWM.read_bytes(), name="sample_iwm_illmitz.xlsx")
    response = auth_client.post(
        _import_url(project),
        {"file": upload, "commit": "true"},
        format="multipart",
    )
    assert response.status_code == 200, response.content

    for ringnummer, sheet_row in fixture.items():
        entry = _capture_for_ringnummer(ringnummer)
        # Ring identity preserved as-is (ADR 0006) — no renumbering.
        assert f"{entry.ring.size}{entry.ring.number}" == ringnummer
        assert entry.species.special_kind  # a Sonderart row, never "unknown species"

        if sheet_row["Art"] == "Ring Vernichtet":
            assert entry.species.special_kind == "ring_destroyed"
            # Bird data nulled server-side despite the sheet carrying a Netz value.
            assert entry.sex is None
            assert entry.age_class is None
            assert entry.bird_status is None
            assert entry.fat_deposit is None
            assert entry.muscle_class is None
            assert entry.net_location is None
        else:
            assert entry.species.special_kind == "unknown_species"
            # Its mandatory Bemerkung and its authentic category codes land intact.
            assert entry.comment == sheet_row["Bemerkungen"]
            assert entry.fat_deposit == int(sheet_row["Fett"])
            assert entry.muscle_class == int(sheet_row["Muskel"])
            assert entry.small_feather_int == int(sheet_row["Intensität"])
            assert entry.hand_wing == int(sheet_row["Handschwingen"])
            assert entry.net_location == int(sheet_row["Netz"])


# --- Biometrics + Fortschritt (issue #176) ------------------------------------
# The export writes the decimal biometrics (Flügellänge/Teilfederlänge/Gewicht/
# Tarsus) and the moult Fortschritt (small_feather_app), but the importer used to
# drop them — so export→import was not a round-trip. A lenient decimal parser now
# reads both point- and comma-separated cells (German-locale files), ignores a
# single garbage cell without blocking the row, and the columns stay optional.

# The optional biometric + Fortschritt columns, layered onto the spine headers.
BIOMETRIC_HEADERS = [*HEADERS, "Flügellänge", "Teilfederlänge", "Gewicht", "Tarsus", "Fortschritt"]


@pytest.mark.django_db
def test_commit_imports_biometrics_with_point_separator(
    auth_client, scientist, ringing_station, project, species
):
    content = _workbook(
        [
            _valid_row(
                species,
                scientist,
                ringing_station,
                **{
                    "Flügellänge": "82.5",
                    "Teilfederlänge": "63.0",
                    "Gewicht": "18.3",
                    "Tarsus": "20.1",
                },
            )
        ],
        headers=BIOMETRIC_HEADERS,
    )

    response = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )

    assert response.status_code == 200, response.content
    assert response.json()["created"] == 1
    entry = DataEntry.objects.get()
    assert entry.wing_span == Decimal("82.50")
    assert entry.feather_span == Decimal("63.00")
    assert entry.weight_gram == Decimal("18.30")
    assert entry.tarsus == Decimal("20.10")


@pytest.mark.django_db
def test_commit_imports_biometrics_with_comma_separator(
    auth_client, scientist, ringing_station, project, species
):
    # German-locale Datenmeldung files write the decimal comma ("18,3"); it must
    # import to the same value a point-separated cell would.
    content = _workbook(
        [
            _valid_row(
                species,
                scientist,
                ringing_station,
                **{
                    "Flügellänge": "82,5",
                    "Teilfederlänge": "63,0",
                    "Gewicht": "18,3",
                    "Tarsus": "20,1",
                },
            )
        ],
        headers=BIOMETRIC_HEADERS,
    )

    response = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )

    assert response.status_code == 200, response.content
    entry = DataEntry.objects.get()
    assert entry.wing_span == Decimal("82.50")
    assert entry.feather_span == Decimal("63.00")
    assert entry.weight_gram == Decimal("18.30")
    assert entry.tarsus == Decimal("20.10")


@pytest.mark.django_db
def test_commit_imports_fortschritt_onto_capture(
    auth_client, scientist, ringing_station, project, species
):
    # The moult Kleingefieder-Fortschritt is a text code (J/U/M/N) the export
    # writes and the import must read back onto small_feather_app.
    content = _workbook(
        [_valid_row(species, scientist, ringing_station, Fortschritt="M")],
        headers=BIOMETRIC_HEADERS,
    )

    response = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )

    assert response.status_code == 200, response.content
    entry = DataEntry.objects.get()
    assert entry.small_feather_app == DataEntry.SmallFeatherAppMoult.MIXED


@pytest.mark.django_db
def test_single_garbage_biometric_cell_is_ignored_and_does_not_block_row(
    auth_client, scientist, ringing_station, project, species
):
    # One unparseable biometric cell is left empty (lenient, like the integer
    # parser) — it never turns into a blocking error, and the good sibling cells
    # on the same row still import.
    content = _workbook(
        [
            _valid_row(
                species,
                scientist,
                ringing_station,
                **{"Flügellänge": "keine Ahnung", "Gewicht": "18.3"},
            )
        ],
        headers=BIOMETRIC_HEADERS,
    )

    response = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )

    assert response.status_code == 200, response.content
    body = response.json()
    assert body["created"] == 1
    assert body["errors"] == []
    entry = DataEntry.objects.get()
    assert entry.wing_span is None  # the garbage cell is ignored, left empty
    assert entry.weight_gram == Decimal("18.30")  # the good cell still imports


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("12.5", Decimal("12.50")),  # point separator
        ("12,5", Decimal("12.50")),  # comma separator (German locale)
        (12.5, Decimal("12.50")),  # openpyxl float
        (12, Decimal("12.00")),  # openpyxl int
        ("", None),  # blank string
        (None, None),  # empty cell
        ("junk", None),  # unparseable
    ],
)
def test_parse_decimal_is_lenient(raw, expected):
    # Direct unit coverage of the lenient decimal parser (issue #176 AC): numerics
    # and strings, both separators, quantized to two places; blank/junk → None.
    assert _parse_decimal(raw) == expected


# --- The "Ring" column: foreign Zentralen, clear rejections, round-trip -------
# (issue #231, PRD #226). The importer reads the "Ring" column to attribute each
# row to its issuing Zentrale, fixing the silent corruption where a foreign ring
# imported as an Austrian one. The Zentrale/Ringgröße rules are the shared
# backend-owned rules of the capture write path (#229) — no divergent logic.
#
# - "Ring" absent or blank → AUW (old sheets stay unbroken).
# - "Ring" = AUW → strict Austrian parsing, exactly as before.
# - "Ring" = a known foreign scheme code → the ring is created under that
#   Zentrale, the Ringnummer split by the generic letters+digits regex into
#   free-text Größe + Nummer.
# - an unsplittable foreign Ringnummer or an unknown scheme code rejects the row
#   with a clear German message — never a silent mis-import.


@pytest.mark.django_db
def test_known_foreign_scheme_imports_under_that_zentrale_with_split_ringnummer(
    auth_client, scientist, ringing_station, project, species, organization
):
    # A Wiederfang of a Slovak-ringed bird: Ring="SKB", Ringnummer "S1234". The row
    # imports under the Slovak Bratislava Zentrale, and the Ringnummer is split by
    # the generic letters+digits regex into free-text Größe "S" + Nummer "1234". (US 22)
    content = _workbook(
        [
            _valid_row(
                species,
                scientist,
                ringing_station,
                Ring="SKB",
                Ringnummer="S1234",
                Ringstatus="W",
            )
        ]
    )

    response = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )

    assert response.status_code == 200, response.content
    assert response.json()["created"] == 1
    entry = DataEntry.objects.get()
    assert entry.ring.central.scheme_code == "SKB"
    assert entry.ring.size == "S"
    assert entry.ring.number == "1234"
    assert entry.ring.organization == organization


@pytest.mark.django_db
def test_sheet_without_ring_column_imports_as_auw(
    auth_client, scientist, ringing_station, project, species
):
    # An old sheet that predates the "Ring" column: the importer defaults every
    # row to the domestic AUW Zentrale, so existing import workflows stay unbroken. (US 25)
    headers = [h for h in HEADERS if h != "Ring"]
    content = _workbook([_valid_row(species, scientist, ringing_station)], headers=headers)

    response = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )

    assert response.status_code == 200, response.content
    assert response.json()["created"] == 1
    entry = DataEntry.objects.get()
    assert entry.ring.central.scheme_code == "AUW"
    assert entry.ring.size == "V"
    assert entry.ring.number == "00604"


@pytest.mark.django_db
def test_blank_ring_cell_imports_as_auw(auth_client, scientist, ringing_station, project, species):
    # The "Ring" column is present but the cell is blank: still AUW, just like an
    # absent column (US 25).
    content = _workbook([_valid_row(species, scientist, ringing_station, Ring="")])

    response = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )

    assert response.status_code == 200, response.content
    assert response.json()["created"] == 1
    entry = DataEntry.objects.get()
    assert entry.ring.central.scheme_code == "AUW"


@pytest.mark.django_db
def test_auw_row_keeps_strict_austrian_size_parsing(
    auth_client, scientist, ringing_station, project, species
):
    # An explicit Ring="AUW" keeps the strict Austrian parse: a Ringnummer whose
    # leading letters are not one of the 28 Austrian codes is rejected exactly as
    # before the Zentrale slice — free text is only for foreign Zentralen. (US 22)
    content = _workbook(
        [_valid_row(species, scientist, ringing_station, Ring="AUW", Ringnummer="ZZ123")]
    )

    preview = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    ).json()

    assert preview["importable"] == 0
    assert [e["row"] for e in preview["errors"]] == [2]
    assert "Ungültige Ringnummer" in preview["errors"][0]["reason"]
    assert DataEntry.objects.count() == 0


@pytest.mark.django_db
def test_unknown_scheme_code_rejects_the_row_with_a_clear_message(
    auth_client, scientist, ringing_station, project, species
):
    # A typo in the "Ring" column (ZZZ is not a seeded EURING scheme) must surface
    # as a blocking error instead of corrupting data — reported already in the
    # dry-run preview and skipped on commit, alongside a good AUW row. (US 24)
    rows = [
        _valid_row(species, scientist, ringing_station),  # row 2 — AUW, importable
        _valid_row(
            species, scientist, ringing_station, Ring="ZZZ", Ringnummer="V00777", Ringstatus="W"
        ),  # row 3 — unknown scheme code
    ]
    content = _workbook(rows)

    preview = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    ).json()

    assert preview["importable"] == 1
    assert [e["row"] for e in preview["errors"]] == [3]
    assert "Unbekannter Zentralen-Code" in preview["errors"][0]["reason"]
    assert "ZZZ" in preview["errors"][0]["reason"]
    assert DataEntry.objects.count() == 0

    result = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    ).json()
    assert result["created"] == 1
    assert [e["row"] for e in result["errors"]] == [3]
    assert "Unbekannter Zentralen-Code" in result["errors"][0]["reason"]
    assert DataEntry.objects.count() == 1


@pytest.mark.django_db
def test_unsplittable_foreign_ringnummer_rejects_with_manual_entry_message(
    auth_client, scientist, ringing_station, project, species
):
    # A foreign Ringnummer the generic letters+digits regex cannot split (an exotic
    # hyphenated format) is rejected with a German message telling the Admin to
    # record it manually — never a silent mis-import. Reported in preview and
    # skipped on commit. (US 23)
    rows = [
        _valid_row(species, scientist, ringing_station),  # row 2 — AUW, importable
        _valid_row(
            species,
            scientist,
            ringing_station,
            Ring="HGB",
            Ringnummer="AB-12345",
            Ringstatus="W",
        ),  # row 3 — unsplittable foreign Ringnummer
    ]
    content = _workbook(rows)

    preview = auth_client.post(
        _import_url(project), {"file": _upload(content)}, format="multipart"
    ).json()

    assert preview["importable"] == 1
    assert [e["row"] for e in preview["errors"]] == [3]
    reason = preview["errors"][0]["reason"]
    assert "AB-12345" in reason
    assert "manuell" in reason  # points the Admin at manual entry
    assert DataEntry.objects.count() == 0

    result = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    ).json()
    assert result["created"] == 1
    assert [e["row"] for e in result["errors"]] == [3]
    assert "manuell" in result["errors"][0]["reason"]
    assert DataEntry.objects.count() == 1


@pytest.mark.django_db
def test_slovak_s_ring_lands_under_skb_and_coexists_with_austrian_s(
    auth_client, scientist, ringing_station, project, species, organization
):
    # The bug this slice fixes: a Slovak "S 1234" used to import silently as an
    # Austrian S ring (both use the letter "S"). Now the "Ring" column routes each
    # to its own Zentrale, so an Austrian S 1234 Erstfang and a Slovak S 1234
    # Wiederfang are two distinct physical rings that coexist without colliding
    # (US 18). The datetimes differ so neither is a duplicate of the other.
    rows = [
        _valid_row(
            species,
            scientist,
            ringing_station,
            Ring="AUW",
            Ringnummer="S1234",
            Ringstatus="E",
            Uhrzeit=time(8, 15),
        ),
        _valid_row(
            species,
            scientist,
            ringing_station,
            Ring="SKB",
            Ringnummer="S1234",
            Ringstatus="W",
            Uhrzeit=time(9, 30),
        ),
    ]
    content = _workbook(rows)

    response = auth_client.post(
        _import_url(project), {"file": _upload(content), "commit": "true"}, format="multipart"
    )

    assert response.status_code == 200, response.content
    assert response.json()["created"] == 2

    rings = Ring.objects.filter(size="S", number="1234", organization=organization)
    assert rings.count() == 2  # distinct physical rings, keyed per Zentrale
    schemes = {r.central.scheme_code for r in rings}
    assert schemes == {"AUW", "SKB"}


@pytest.mark.django_db
def test_reimporting_our_own_export_with_foreign_rows_round_trips(
    auth_client, scientist, ringing_station, project, species, organization
):
    # Export → import is a faithful loop even for the foreign rows #230 emits: a
    # domestic AUW Erstfang and a Slovak (SKB) Wiederfang exported to a
    # Datenmeldung and re-imported reconstruct equivalent captures, each under its
    # own Zentrale. (US 26)
    ringing_station.place_code = "AU03"
    ringing_station.region = "Oberösterreich"
    ringing_station.country = "Austria"
    ringing_station.save()

    skb = Central.objects.get(scheme_code="SKB")
    auw = Central.objects.get(scheme_code="AUW")

    domestic_ring = Ring.objects.create(
        number="00604", size="V", organization=organization, central=auw
    )
    DataEntry.objects.create(
        species=species,
        ring=domestic_ring,
        staff=scientist,
        ringing_station=ringing_station,
        project=project,
        organization=organization,
        bird_status=DataEntry.BirdStatus.FIRST_CATCH,
        date_time=make_aware(datetime(2026, 6, 30, 8, 15)),
    )
    foreign_ring = Ring.objects.create(
        number="1234", size="S", organization=organization, central=skb
    )
    DataEntry.objects.create(
        species=species,
        ring=foreign_ring,
        staff=scientist,
        ringing_station=ringing_station,
        project=project,
        organization=organization,
        bird_status=DataEntry.BirdStatus.RE_CATCH,
        date_time=make_aware(datetime(2026, 7, 15, 9, 0)),
    )

    content = build_iwm_workbook(DataEntry.objects.filter(project=project).order_by("date_time"))

    # Remove the source captures so the re-import is not skipped as a duplicate of
    # the very data it round-trips, then re-import into a fresh Projekt in the same
    # Organisation (so the Beringer/Station re-resolve, no auto-creation).
    DataEntry.objects.filter(project=project).delete()
    fresh_project = Project.objects.create(title="Round-Trip", organization=organization)

    result = commit_import(content, fresh_project)

    assert result["created"] == 2
    assert result["errors"] == []
    assert result["createdBeringer"] == []
    assert result["createdStationen"] == []

    clones = {
        (e.ring.central.scheme_code, e.ring.size, e.ring.number)
        for e in DataEntry.objects.filter(project=fresh_project)
    }
    assert clones == {("AUW", "V", "00604"), ("SKB", "S", "1234")}
