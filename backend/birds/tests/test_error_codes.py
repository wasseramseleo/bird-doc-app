"""Die Domänencodes: every hand-written rejection names its own cause.

The previous slice gave every rejection an ``errors`` envelope carrying DRF's own
generic codes (``required``, ``blank``, ``invalid_choice``, …). Those say *which
condition* tripped, never *which rule* — a doubly-issued Ringnummer and a
mistyped decimal both arrived as ``invalid``. This slice names the causes, in one
pass, so a client can act on them (ADR 0038).

Every code asserted here is spelled out as a **literal**, never as an imported
constant: a published code is a **contract**, so a rename has to break a test
rather than be carried along by it. The German sentences are asserted through
``_without_envelope`` exactly as in ``test_error_envelope.py`` — the ADR 0033
byte-identity constraint holds unchanged, this slice only adds meaning to the
sibling key.

Driven through the existing DRF HTTP API with the fixtures from ``conftest.py``.
The one rule no request can reach — a blank foreign Ringgröße, refused as a
field-level ``blank`` long before the service sees it — is exercised at the
capture service, the seam ``test_capture_service.py`` already uses, and the
*reason* it cannot reach the wire is pinned through the API beside it.
"""

from datetime import date, time
from io import BytesIO

import openpyxl
import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework.test import APIClient

from birds.capture_service import CaptureValidationError, create_capture
from birds.models import Central, DataEntry, Mitgliedschaft
from birds.payload_schema import UnmigratablePayloadError

DATA_ENTRIES_URL = "/api/birds/data-entries/"
FEEDBACK_URL = "/api/feedback/"
INVITATIONS_URL = "/api/birds/invitations/"
LOGIN_URL = "/api/auth/login/"
ME_URL = "/api/auth/me/"
MITGLIEDSCHAFTEN_URL = "/api/birds/mitgliedschaften/"
PROJECTS_URL = "/api/birds/projects/"
SCIENTISTS_URL = "/api/birds/scientists/"
STATIONS_URL = "/api/birds/ringing-stations/"


def _payload(species, scientist, ringing_station, *, ring_number="200", ring_size="V"):
    return {
        "species_id": str(species.id),
        "staff_id": scientist.id,
        "ringing_station_id": ringing_station.handle,
        "ring_number": ring_number,
        "ring_size": ring_size,
        "date_time": "2026-03-01T12:00:00Z",
    }


def _without_envelope(body):
    """The body an old bundle sees: everything but the sibling key."""
    return {key: value for key, value in body.items() if key != "errors"}


def _codes(body):
    return [entry["code"] for entry in body["errors"]]


def _minimal_fangdaten_upload():
    """A structurally valid one-row ``Datenmeldung``, so the row cap is what
    refuses it rather than the structure check that runs first."""
    from birds.iwm_import import REQUIRED_HEADERS

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Fangdaten"
    headers = sorted(REQUIRED_HEADERS)
    sheet.append(headers)
    sheet.append(["x"] * len(headers))
    buffer = BytesIO()
    workbook.save(buffer)
    return SimpleUploadedFile(
        "import.xlsx",
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# The authentic ``Fangdaten`` columns a row-level test needs (a superset of
# ``REQUIRED_HEADERS``; the importer reads by name), mirroring the layout
# ``test_iwm_import.py`` builds its sheets from.
_FANGDATEN_HEADERS = [
    "Ring",
    "Ringnummer",
    "Ringstatus",
    "Art",
    "Geschlecht",
    "Alter",
    "Datum",
    "Uhrzeit",
    "Ortskodierung",
    "Ort",
    "Bemerkungen",
    "BeringerIn",
]


def _fangdaten_row(species, scientist, ringing_station, **overrides):
    """A structurally complete, resolvable row for tenant A."""
    row = {
        "Ring": "AUW",
        "Ringnummer": "V00604",
        "Ringstatus": "E",
        "Art": species.common_name_de,
        "Geschlecht": "U",
        "Alter": 3,
        "Datum": date(2026, 6, 30),
        "Uhrzeit": time(8, 15),
        "Ort": ringing_station.name,
        "Bemerkungen": "",
        "BeringerIn": scientist.handle,
    }
    row.update(overrides)
    return row


def _fangdaten_upload(rows):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Fangdaten"
    sheet.append(_FANGDATEN_HEADERS)
    columns = {header: i + 1 for i, header in enumerate(_FANGDATEN_HEADERS)}
    for row_offset, row in enumerate(rows, start=2):
        for header, column in columns.items():
            value = row.get(header)
            if value not in (None, ""):
                sheet.cell(row=row_offset, column=column, value=value)
    buffer = BytesIO()
    workbook.save(buffer)
    return SimpleUploadedFile(
        "import.xlsx",
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _row_codes(report):
    """The per-row rejections of an ``ImportPreview``/``ImportResult``, keyed by
    sheet row: the Import-Dialog's own list, which carries its own ``code``
    beside the German ``reason`` (the envelope is for the *whole-file* refusals)."""
    return {entry["row"]: entry["code"] for entry in report["errors"]}


# --- Capture-Pfad (ADR 0004, 0006, 0019, 0026) -------------------------------


@pytest.mark.django_db
def test_ring_already_first_caught_names_its_own_rule(
    auth_client, species, scientist, ringing_station
):
    """The rejection the whole PRD started from stops arriving as a bare
    ``invalid``: the client can now offer „Als Wiederfang" off the code alone."""
    first = _payload(species, scientist, ringing_station, ring_number="901")
    first["bird_status"] = "e"
    assert auth_client.post(DATA_ENTRIES_URL, first, format="json").status_code == 201

    again = _payload(species, scientist, ringing_station, ring_number="901")
    again["bird_status"] = "e"
    response = auth_client.post(DATA_ENTRIES_URL, again, format="json")

    assert response.status_code == 400
    body = response.json()
    assert _without_envelope(body) == {
        "ring_number": "Für diese Ringnummer besteht in dieser Organisation bereits ein Erstfang."
    }
    assert body["errors"] == [
        {
            "field": "ring_number",
            "code": "ring_already_first_caught",
            "detail": ("Für diese Ringnummer besteht in dieser Organisation bereits ein Erstfang."),
        }
    ]


@pytest.mark.django_db
def test_invalid_austrian_ring_size_names_its_own_rule(
    auth_client, species, scientist, ringing_station
):
    """Under AUW the strict Austrian choice list governs (ADR 0019) — a distinct
    cause from the free-text rule below, so a distinct code."""
    payload = _payload(species, scientist, ringing_station, ring_number="902", ring_size="ZZ")

    response = auth_client.post(DATA_ENTRIES_URL, payload, format="json")

    assert response.status_code == 400
    body = response.json()
    assert _without_envelope(body) == {"ring_size": "Keine gültige österreichische Ringgröße."}
    assert _codes(body) == ["ring_size_invalid_austrian"]


@pytest.mark.django_db
def test_foreign_ring_size_required_guards_the_service_seam_not_the_wire(
    species, scientist, ringing_station, organization
):
    """The other Ringgröße rule, and the one code in the catalogue **no request
    can observe** — noted as such at its entry in ``error_codes.py``.

    No caller can hand ``normalize_ring_size`` a blank Größe: the DRF field
    refuses one first (pinned by the test below) and the IWM importer's
    letters+digits split cannot produce one (pinned by
    ``test_iwm_unsplittable_foreign_ringnummer_never_reaches_a_blank_groesse``).
    So it is asserted where it actually bites — the shared service seam, whose
    contract is with the *next* caller, not with a client."""
    with pytest.raises(CaptureValidationError) as exc_info:
        create_capture(
            species=species,
            staff=scientist,
            ringing_station=ringing_station,
            organization=organization,
            ring_size="   ",
            ring_number="SK1",
            date_time="2026-03-01T12:00:00Z",
            central=Central.objects.get(scheme_code="SKB"),
            bird_status=DataEntry.BirdStatus.RE_CATCH,
        )

    assert exc_info.value.field == "ring_size"
    assert exc_info.value.code == "ring_size_required_foreign"


@pytest.mark.django_db
def test_a_blank_ring_size_is_refused_at_the_field_long_before_the_rule_speaks(
    auth_client, species, scientist, ringing_station
):
    """Why the rule above cannot reach the wire, pinned rather than asserted in
    prose. ``DataEntrySerializer.ring_size`` is a plain ``CharField`` (no
    ``allow_blank``, whitespace trimmed), so ``""``/``"   "`` is DRF's generic
    ``blank`` — the condition, correctly — and ``normalize_ring_size`` never runs.
    Relaxing that field would make the catalogue's note false, so it breaks this
    test first."""
    payload = _payload(species, scientist, ringing_station, ring_number="907", ring_size="   ")
    payload["central"] = "SKB"
    payload["bird_status"] = "w"

    response = auth_client.post(DATA_ENTRIES_URL, payload, format="json")

    assert response.status_code == 400
    body = response.json()
    assert list(_without_envelope(body)) == ["ring_size"]
    assert _codes(body) == ["blank"]


@pytest.mark.django_db
def test_status_requires_projekt_zentrale_names_its_own_rule(
    auth_client, species, scientist, ringing_station
):
    """An Erstfang draws a fresh number from the Projekt's own rope, so it may
    only be issued under the Projekt-Zentrale (ADR 0019). The rejection names no
    single field the Beringer typed — it is the *combination* that is refused."""
    payload = _payload(species, scientist, ringing_station, ring_number="903")
    payload["central"] = "SKB"
    payload["bird_status"] = "e"

    response = auth_client.post(DATA_ENTRIES_URL, payload, format="json")

    assert response.status_code == 400
    body = response.json()
    assert _without_envelope(body) == {
        "central": (
            "Erstfänge und vernichtete Ringe müssen unter der Projekt-Zentrale erfasst werden."
        )
    }
    assert _codes(body) == ["status_requires_projekt_zentrale"]


@pytest.mark.django_db
def test_aves_ignota_comment_requirement_names_its_own_rule(
    auth_client, aves_ignota_species, scientist, ringing_station
):
    """The unusual catch must always be described (ADR 0004)."""
    payload = _payload(aves_ignota_species, scientist, ringing_station, ring_number="904")

    response = auth_client.post(DATA_ENTRIES_URL, payload, format="json")

    assert response.status_code == 400
    body = response.json()
    assert _without_envelope(body) == {
        "comment": ["Für eine unbekannte Art (Aves ignota) ist eine Bemerkung erforderlich."]
    }
    assert _codes(body) == ["aves_ignota_comment_required"]


@pytest.mark.django_db
def test_marker_comment_requirement_names_its_own_rule(
    auth_client, species, scientist, ringing_station
):
    """A Tot-Fund or Nicht-Standard-Fang must always be described (ADR 0026) —
    the same field as Aves ignota, a different cause, therefore a different
    code."""
    payload = _payload(species, scientist, ringing_station, ring_number="905")
    payload["is_dead_recovery"] = True

    response = auth_client.post(DATA_ENTRIES_URL, payload, format="json")

    assert response.status_code == 400
    body = response.json()
    assert _without_envelope(body) == {
        "comment": ["Bei einem Tot-Fund oder Nicht-Standard-Fang ist eine Bemerkung erforderlich."]
    }
    assert _codes(body) == ["marker_comment_required"]


@pytest.mark.django_db
def test_unknown_zentralen_code_names_its_own_rule(
    auth_client, species, scientist, ringing_station
):
    """One of the five ``error_messages`` dictionaries: the hand-written German
    sentence replaces DRF's ``does_not_exist`` prose, and the code now names the
    cause instead of the generic condition."""
    payload = _payload(species, scientist, ringing_station, ring_number="906")
    payload["central"] = "ZZZ"
    payload["bird_status"] = "w"

    response = auth_client.post(DATA_ENTRIES_URL, payload, format="json")

    assert response.status_code == 400
    body = response.json()
    assert _without_envelope(body) == {"central": ["Unbekannter Zentralen-Code."]}
    assert _codes(body) == ["central_unknown"]


# --- 403 disambiguiert (ADR 0037) --------------------------------------------
# A Rechteverweigerung and a CSRF-Ablehnung are the same status with opposite
# ways out: the first needs another person, the second just needs pressing again.


@pytest.mark.django_db
def test_rechteverweigerung_carries_admin_only(mitglied_client, mitglied_scientist):
    """A Mitglied refused an Admin-only action: the way out is another person."""
    response = mitglied_client.post(
        "/api/birds/ringing-stations/", {"name": "Neue Station"}, format="json"
    )

    assert response.status_code == 403
    body = response.json()
    assert _codes(body) == ["admin_only"]


@pytest.mark.django_db
def test_csrf_ablehnung_carries_a_different_code_than_rechteverweigerung(
    db, mitglied_user, mitglied_scientist
):
    """The same 403, the opposite way out — pressing again is enough and no Admin
    can help. It comes out of DRF's ``SessionAuthentication``, not out of a raise
    of ours, so it is coded where it is born rather than recognised by its prose.
    """
    client = APIClient(enforce_csrf_checks=True)
    assert client.login(username="mara", password="hunter2-very-strong")

    response = client.post(DATA_ENTRIES_URL, {}, format="json")

    assert response.status_code == 403
    body = response.json()
    assert _codes(body) == ["csrf_failed"]
    # The English DRF sentence is untouched — only the code is new.
    assert _without_envelope(body)["detail"].startswith("CSRF Failed")


@pytest.mark.django_db
def test_no_active_organisation_is_distinguishable_from_admin_only(auth_client, user):
    """Both are a 403 and both need someone else, but the ways out differ: a
    Mitgliedschaft has to be granted, not this one action approved. An account
    with no membership at all cannot attach a row to a tenant (ADR 0005)."""
    response = auth_client.post("/api/birds/scientists/", {"handle": "NEU"}, format="json")

    assert response.status_code == 403
    body = response.json()
    assert _without_envelope(body) == {
        "detail": "Keine aktive Organisation — eine Mitgliedschaft ist erforderlich."
    }
    assert _codes(body) == ["no_active_organisation"]


@pytest.mark.django_db
def test_managing_another_organisations_station_carries_its_own_code(
    auth_client, scientist, organization_b, ringing_station
):
    """An Admin may only manage their own Organisation's rows — a third distinct
    403, and the one nobody can grant their way out of (ADR 0005)."""
    response = auth_client.patch(
        f"/api/birds/ringing-stations/{ringing_station.handle}/",
        {"organization_id": organization_b.pk},
        format="json",
    )

    assert response.status_code == 403
    body = response.json()
    assert _without_envelope(body) == {
        "detail": "Du kannst nur Objekte deiner eigenen Organisation verwalten."
    }
    assert _codes(body) == ["not_own_organisation"]


# --- Mitgliedschaft und Einladung (issue #83) --------------------------------


@pytest.mark.django_db
def test_demoting_the_last_admin_carries_last_admin(auth_client, scientist):
    """Every Organisation keeps at least one Admin — the Rolle-Wechsel path, whose
    body is a mapping and therefore has room for the envelope."""
    membership = Mitgliedschaft.objects.get(user=scientist.user)

    response = auth_client.patch(
        f"{MITGLIEDSCHAFTEN_URL}{membership.pk}/",
        {"rolle": Mitgliedschaft.Rolle.MITGLIED},
        format="json",
    )

    assert response.status_code == 400
    body = response.json()
    assert _codes(body) == ["last_admin"]


@pytest.mark.django_db
def test_removing_the_last_admin_keeps_its_list_body_and_publishes_no_code(auth_client, scientist):
    """The one rejection whose code cannot reach the wire. It is raised with a
    bare sentence, which DRF renders as a JSON **list** — there is no sibling key
    to hang an envelope on, and reshaping the body to make room would break the
    ADR 0033 byte identity a month-old bundle replays into. The German sentence
    still travels, which is the fallback ADR 0038 prescribes for a code a client
    does not have; the raise itself is coded ``last_admin`` all the same, so the
    day the body can carry it, it will."""
    membership = Mitgliedschaft.objects.get(user=scientist.user)

    response = auth_client.delete(f"{MITGLIEDSCHAFTEN_URL}{membership.pk}/")

    assert response.status_code == 400
    assert response.json() == [
        "Die Organisation braucht mindestens eine:n Administrator:in. "
        "Ernenne zuerst eine andere Person zur Administratorin oder zum Administrator."
    ]


@pytest.mark.django_db
def test_seat_limit_and_duplicate_invitation_carry_distinct_codes(
    auth_client, user, organization, mailoutbox
):
    """Three refusals on one endpoint, three causes: a seat already taken by a
    Mitglied, one reserved by a pending Einladung, and the Seat-Limit itself."""
    Mitgliedschaft.objects.create(
        user=user, organization=organization, rolle=Mitgliedschaft.Rolle.ADMIN
    )

    first = auth_client.post(INVITATIONS_URL, {"email": "kollege@example.org"}, format="json")
    assert first.status_code == 201, first.json()

    duplicate = auth_client.post(INVITATIONS_URL, {"email": "kollege@example.org"}, format="json")
    assert duplicate.status_code == 400
    assert _without_envelope(duplicate.json()) == {
        "email": "Für diese E-Mail gibt es bereits eine offene Einladung."
    }
    assert _codes(duplicate.json()) == ["invitation_already_pending"]

    user.email = "alice@example.org"
    user.save(update_fields=["email"])
    already_member = auth_client.post(
        INVITATIONS_URL, {"email": "alice@example.org"}, format="json"
    )
    assert already_member.status_code == 400
    assert _codes(already_member.json()) == ["already_member"]

    organization.seat_limit = 1
    organization.save(update_fields=["seat_limit"])
    over_limit = auth_client.post(INVITATIONS_URL, {"email": "zweiter@example.org"}, format="json")
    assert over_limit.status_code == 409
    assert _codes(over_limit.json()) == ["seat_limit_reached"]


# --- Beringer und Station (PRD #205, issue #117) ------------------------------


@pytest.mark.django_db
def test_re_pointing_a_beringer_with_captures_carries_its_own_code(
    auth_client, scientist, data_entry, gap_seat
):
    """Freeze-once-captures: a linked Beringer that owns Fänge is neither
    detached nor re-pointed, so the capture history stays attributable."""
    response = auth_client.patch(
        f"{SCIENTISTS_URL}{scientist.id}/",
        {"mitgliedschaft_id": str(gap_seat.id)},
        format="json",
    )

    assert response.status_code == 400
    assert _codes(response.json()) == ["beringer_frozen_by_captures"]


@pytest.mark.django_db
def test_linking_a_seat_whose_account_is_already_a_beringer_carries_its_own_code(
    auth_client, scientist, no_account_beringer, mitglied_scientist
):
    """The OneToOne ``Scientist.user`` must be free to attach — a different cause
    on the same field, therefore a different code."""
    taken_seat = Mitgliedschaft.objects.get(user=mitglied_scientist.user)

    response = auth_client.patch(
        f"{SCIENTISTS_URL}{no_account_beringer.id}/",
        {"mitgliedschaft_id": str(taken_seat.id)},
        format="json",
    )

    assert response.status_code == 400
    assert _codes(response.json()) == ["account_already_beringer"]


@pytest.mark.django_db
def test_cross_tenant_seat_link_carries_its_own_code(
    auth_client, scientist, no_account_beringer, scientist_b
):
    """Linking never crosses the tenant boundary (ADR 0005)."""
    foreign_seat = Mitgliedschaft.objects.get(user=scientist_b.user)

    response = auth_client.patch(
        f"{SCIENTISTS_URL}{no_account_beringer.id}/",
        {"mitgliedschaft_id": str(foreign_seat.id)},
        format="json",
    )

    assert response.status_code == 400
    assert _codes(response.json()) == ["mitgliedschaft_other_organisation"]


@pytest.mark.django_db
def test_station_with_captures_cannot_be_deleted_and_says_why(
    auth_client, scientist, data_entry, ringing_station
):
    """A 409 the Admin can act on: archive the Station instead (issue #117)."""
    response = auth_client.delete(f"{STATIONS_URL}{ringing_station.handle}/")

    assert response.status_code == 409
    body = response.json()
    assert _without_envelope(body) == {
        "detail": (
            "Diese Station kann nicht gelöscht werden, weil ihr Fänge zugeordnet sind. "
            "Archiviere die Station stattdessen."
        )
    }
    assert _codes(body) == ["station_has_captures"]


@pytest.mark.django_db
def test_deleting_a_linked_beringer_says_why(auth_client, scientist, mitglied_scientist):
    """A Mitglied is never stripped of their Beringer identity from this screen
    (PRD #205, issue #208) — the way out is member management, so the code names
    the link rather than the deletion."""
    response = auth_client.delete(f"{SCIENTISTS_URL}{mitglied_scientist.id}/")

    assert response.status_code == 409
    body = response.json()
    assert _codes(body) == ["beringer_has_account"]


# --- Projekt (issue #389, #74) ------------------------------------------------


@pytest.mark.django_db
def test_projekt_without_a_beringer_carries_its_own_code(auth_client, membership):
    """Ein Projekt braucht mindestens eine:n Beringer:in — an Admin without a
    Beringer of their own has nothing to auto-add."""
    response = auth_client.post(PROJECTS_URL, {"title": "Ohne Beringer"}, format="json")

    assert response.status_code == 400
    body = response.json()
    assert _without_envelope(body) == {
        "scientist_ids": ["Ein Projekt braucht mindestens eine:n Beringer:in."]
    }
    assert _codes(body) == ["project_requires_scientist"]


@pytest.mark.django_db
def test_projekt_default_station_from_another_organisation_carries_its_own_code(
    auth_client, scientist, ringing_station_b
):
    """The Standard-Station must belong to the Projekt's Organisation (issue #74)."""
    response = auth_client.post(
        PROJECTS_URL,
        {"title": "Fremde Station", "default_station_id": ringing_station_b.handle},
        format="json",
    )

    assert response.status_code == 400
    body = response.json()
    assert _codes(body) == ["station_other_organisation"]


# --- IWM-Import (ADR 0013, issue #125) ---------------------------------------


@pytest.mark.django_db
def test_import_without_a_file_carries_its_own_code(auth_client, scientist, project):
    response = auth_client.post(f"{PROJECTS_URL}{project.id}/import-iwm/", {}, format="multipart")

    assert response.status_code == 400
    body = response.json()
    assert _without_envelope(body) == {"file": "Es wurde keine Datei hochgeladen."}
    assert _codes(body) == ["file_required"]


@pytest.mark.django_db
def test_unreadable_import_file_carries_its_own_code(auth_client, scientist, project):
    """Wrong file, no ``Fangdaten`` sheet, missing columns — one cause („this is
    not a readable Datenmeldung"), so one code, whichever sentence it wears."""
    upload = SimpleUploadedFile("nonsense.xlsx", b"not an excel file at all")

    response = auth_client.post(
        f"{PROJECTS_URL}{project.id}/import-iwm/", {"file": upload}, format="multipart"
    )

    assert response.status_code == 400
    body = response.json()
    assert _without_envelope(body) == {
        "file": "Die Datei konnte nicht als Excel-Arbeitsmappe gelesen werden."
    }
    assert _codes(body) == ["iwm_file_unreadable"]


@pytest.mark.django_db
def test_over_cap_import_carries_its_own_code_and_keeps_its_cap_key(
    monkeypatch, auth_client, scientist, ringing_station, project, species
):
    """A rejection the view builds by hand, because its body carries a typed
    ``cap`` object DRF would stringify. It hangs its own envelope instead, and
    ``file``/``cap`` stay exactly as the Import-Dialog already reads them."""
    monkeypatch.setattr("birds.iwm_import.ROW_CAP", 0)

    response = auth_client.post(
        f"{PROJECTS_URL}{project.id}/import-iwm/",
        {"file": _minimal_fangdaten_upload()},
        format="multipart",
    )

    assert response.status_code == 400
    body = response.json()
    assert body["cap"] == {"limit": 0, "exceeded": True}
    assert "aufteilen" in body["file"]
    assert body["errors"] == [
        {"field": "file", "code": "iwm_row_cap_exceeded", "detail": body["file"]}
    ]


# --- IWM-Import: die Zeilenzurückweisungen ------------------------------------
# The three refusals above are whole-file: the upload never becomes a report. A
# *row* rejection is the one an Admin actually acts on — one per bad row, listed
# in the report's own ``errors`` as ``{row, reason, code}``. The ``code`` is
# additive; the Import-Dialog reads ``row``/``reason`` and is unmoved by it.


@pytest.mark.django_db
def test_an_imported_row_publishes_the_same_rule_the_write_path_does(
    auth_client, aves_ignota_species, scientist, ringing_station, project
):
    """A code names a **cause, not a call site**: the Aves-ignota Bemerkung is
    ``aves_ignota_comment_required`` whether a Beringer typed the capture or an
    Admin imported it. ``CaptureValidationError.code`` is what carries it across
    — the same exception both the serializer and the importer catch."""
    upload = _fangdaten_upload(
        [
            _fangdaten_row(
                aves_ignota_species, scientist, ringing_station, Bemerkungen="", Ringstatus="W"
            )
        ]
    )

    response = auth_client.post(
        f"{PROJECTS_URL}{project.id}/import-iwm/", {"file": upload}, format="multipart"
    )

    assert response.status_code == 200
    report = response.json()
    assert report["errors"] == [
        {
            "row": 2,
            "reason": "Für eine unbekannte Art (Aves ignota) ist eine Bemerkung erforderlich.",
            "code": "aves_ignota_comment_required",
        }
    ]


@pytest.mark.django_db
def test_a_committed_row_rejection_carries_its_code_too(
    auth_client, species, scientist, ringing_station, project
):
    """The other seam: a rejection only the *write* raises. The dry-run cannot
    see the Erstfang collision (it writes nothing), so the commit path has its own
    ``CaptureValidationError`` handler — and it must publish the same code the
    online capture write publishes for the very same ring."""
    first = _payload(species, scientist, ringing_station, ring_number="00604")
    first["bird_status"] = "e"
    assert auth_client.post(DATA_ENTRIES_URL, first, format="json").status_code == 201

    upload = _fangdaten_upload([_fangdaten_row(species, scientist, ringing_station)])
    response = auth_client.post(
        f"{PROJECTS_URL}{project.id}/import-iwm/",
        {"file": upload, "commit": "true"},
        format="multipart",
    )

    assert response.status_code == 200
    report = response.json()
    assert report["created"] == 0
    assert report["errors"] == [
        {
            "row": 2,
            "reason": "Für diese Ringnummer besteht in dieser Organisation bereits ein Erstfang.",
            "code": "ring_already_first_caught",
        }
    ]


@pytest.mark.django_db
def test_the_importers_own_row_refusals_each_name_their_cause(
    auth_client, species, scientist, ringing_station, project
):
    """The rest of the row-level surface: refusals the importer writes itself,
    before a capture is ever built. All of them, in one file — a half-coded list
    would leave the Dialog unable to tell „a code I do not know" from „a row error
    with no code" (ADR 0038). The unknown Zentralen-Code reuses the write path's
    ``central_unknown``: same cause, different call site, one code."""
    upload = _fangdaten_upload(
        [
            _fangdaten_row(species, scientist, ringing_station, Ringnummer=""),
            _fangdaten_row(species, scientist, ringing_station, Ringnummer="V99999", Ring="ZZZ"),
            _fangdaten_row(species, scientist, ringing_station, Ringnummer="V-nonsense"),
            _fangdaten_row(species, scientist, ringing_station, Datum=""),
            _fangdaten_row(species, scientist, ringing_station, Datum="übermorgen"),
            _fangdaten_row(species, scientist, ringing_station, Art=""),
            _fangdaten_row(species, scientist, ringing_station, Art="Ferkelvogel"),
            _fangdaten_row(species, scientist, ringing_station, BeringerIn=""),
            _fangdaten_row(species, scientist, ringing_station, Ort="", Ortskodierung=""),
        ]
    )

    response = auth_client.post(
        f"{PROJECTS_URL}{project.id}/import-iwm/", {"file": upload}, format="multipart"
    )

    assert response.status_code == 200
    report = response.json()
    assert _row_codes(report) == {
        2: "iwm_row_ring_number_missing",
        3: "central_unknown",
        4: "iwm_row_ring_number_invalid",
        5: "iwm_row_date_missing",
        6: "iwm_row_date_invalid",
        7: "iwm_row_species_missing",
        8: "iwm_row_species_unknown",
        9: "iwm_row_beringer_missing",
        10: "iwm_row_place_missing",
    }
    # The German prose is untouched — the code rides beside it, never instead.
    assert report["errors"][0]["reason"] == "Ringnummer fehlt."


@pytest.mark.django_db
def test_a_kuerzel_owned_by_another_organisation_names_its_cause(
    auth_client, species, scientist, ringing_station, project, scientist_b
):
    """``Scientist.handle`` is globally unique, so an unfamiliar Kürzel already
    owned elsewhere cannot be auto-created — a blocking row, not a 500. The way
    out is a different Kürzel in the file, which is what the code names."""
    upload = _fangdaten_upload(
        [_fangdaten_row(species, scientist, ringing_station, BeringerIn=scientist_b.handle)]
    )

    response = auth_client.post(
        f"{PROJECTS_URL}{project.id}/import-iwm/", {"file": upload}, format="multipart"
    )

    assert response.status_code == 200
    report = response.json()
    assert _row_codes(report) == {2: "iwm_row_beringer_handle_taken"}


@pytest.mark.django_db
def test_iwm_unsplittable_foreign_ringnummer_never_reaches_a_blank_groesse(
    auth_client, species, scientist, ringing_station, project
):
    """The second half of why ``ring_size_required_foreign`` cannot reach the
    wire. A foreign Ringnummer is split by ``^([A-Za-z]+)(\\d+)$``, so the Größe
    the importer hands the capture service is never empty: a value that would
    yield one is refused *here*, one row earlier, with its own code."""
    upload = _fangdaten_upload(
        [_fangdaten_row(species, scientist, ringing_station, Ring="SKB", Ringnummer="12345")]
    )

    response = auth_client.post(
        f"{PROJECTS_URL}{project.id}/import-iwm/", {"file": upload}, format="multipart"
    )

    assert response.status_code == 200
    report = response.json()
    assert _row_codes(report) == {2: "iwm_row_ring_number_unsplittable"}
    assert "manuell" in report["errors"][0]["reason"]


# --- Anmeldung und Sitzung (ADR 0008) -----------------------------------------


@pytest.mark.django_db
def test_login_refusal_carries_its_own_code(api_client, user):
    """Two doors into one cause — no credentials given, or wrong ones — so one
    code. Built by hand as a plain ``Response``, so it hangs its own envelope."""
    empty = api_client.post(LOGIN_URL, {}, format="json")
    wrong = api_client.post(LOGIN_URL, {"username": "alice", "password": "nope"}, format="json")

    for response in (empty, wrong):
        assert response.status_code == 401
        body = response.json()
        assert _without_envelope(body) == {
            "detail": "Anmeldung fehlgeschlagen. Bitte überprüfe Benutzernamen und Passwort."
        }
        assert _codes(body) == ["login_failed"]


@pytest.mark.django_db
def test_unauthenticated_me_carries_its_own_code(api_client):
    """„Neu anmelden" — a different cause from a failed attempt: the session ended
    rather than the credentials being wrong."""
    response = api_client.get(ME_URL)

    assert response.status_code == 401
    body = response.json()
    assert _without_envelope(body) == {"detail": "Not authenticated."}
    assert _codes(body) == ["not_authenticated"]


# --- Feedback, Ringnummer-Vorschlag, Datumsfilter -----------------------------


@pytest.mark.django_db
def test_empty_feedback_carries_its_own_code(auth_client, user):
    response = auth_client.post("/api/feedback/", {"message": "   "}, format="json")

    assert response.status_code == 400
    body = response.json()
    assert _without_envelope(body) == {"detail": "Bitte gib eine Nachricht ein."}
    assert _codes(body) == ["feedback_message_required"]


@pytest.mark.django_db
def test_next_number_without_a_size_carries_its_own_code(auth_client, membership):
    """The one rejection whose body is keyed ``error`` rather than ``detail`` — an
    old shape kept byte-identical, with the envelope beside it."""
    response = auth_client.get("/api/birds/rings/next-number/")

    assert response.status_code == 400
    body = response.json()
    assert _without_envelope(body) == {"error": "Ring size parameter is required."}
    assert _codes(body) == ["ring_size_parameter_required"]


@pytest.mark.django_db
def test_malformed_date_filter_carries_its_own_code(auth_client, scientist, project):
    response = auth_client.get(f"{PROJECTS_URL}{project.id}/stats/?from=nicht-ein-datum")

    assert response.status_code == 400
    body = response.json()
    assert _codes(body) == ["date_invalid"]


@pytest.mark.django_db
def test_unmigratable_payload_is_accepted_and_still_names_what_happened(
    auth_client, scientist, monkeypatch
):
    """Not a refusal: a payload the server cannot bring forward is **accepted**
    (200) so the device dequeues and nothing strands (ADR 0033). It is held raw
    and the operator alerted — and the envelope names that outcome, so the one
    2xx that is not an ordinary success is still machine-readable."""

    def unmigratable(payload):
        raise UnmigratablePayloadError("2019-01-01")

    monkeypatch.setattr("birds.views.migrate_payload", unmigratable)

    response = auth_client.post(DATA_ENTRIES_URL, {}, format="json")

    assert response.status_code == 200
    body = response.json()
    assert _without_envelope(body) == {
        "detail": (
            "Der Fang wurde übernommen, konnte aber nicht ausgewertet werden und wird geprüft."
        )
    }
    assert _codes(body) == ["unmigratable_payload"]
