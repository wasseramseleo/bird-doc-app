"""Deterministic IWM anonymiser — the real→demo de-identification transform
(issue #177, ADR 0012).

Highest seam: the pure workbook transform ``anonymize_workbook`` exercised over an
in-memory ``Fangdaten`` workbook — the same authentic layout the importer reads.
The synthetic ``sample_iwm_illmitz.xlsx`` doubles as the anonymiser's test input
(it deliberately carries the shape the anonymiser must collapse: 5 Beringer,
3 sub-sites, ring reuse across Erstfang/Wiederfang/Ring-Vernichtet, two Sonderart
kinds). A thin ``call_command('anonymize_iwm', …)`` smoke test covers the CLI
wrapper with temp input/output files.

Every assertion is on externally observable workbook cells (parsed values), never
on private helpers.
"""

import re
from collections import Counter, defaultdict
from datetime import datetime, time
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
from django.core.management import call_command
from django.core.management.base import CommandError

from birds.iwm_anonymize import (
    AVES_IGNOTA_ART,
    AVES_IGNOTA_PLACEHOLDER,
    CURATED_KUERZEL,
    CURATED_STATION,
    DEMO_RING_BASE,
    DEMO_RING_WIDTH,
    TARGET_END_YEAR,
    anonymize_workbook,
)

SAMPLE_IWM = Path(__file__).resolve().parent.parent / "demo" / "sample_iwm_illmitz.xlsx"

# A compact authentic-format header set for the small controlled workbooks.
CTRL_HEADERS = [
    "Ringnummer",
    "Ringstatus",
    "Art",
    "Geschlecht",
    "Alter",
    "Datum",
    "Uhrzeit",
    "Ortskodierung",
    "Geo-Koordinaten",
    "Ort",
    "Region",
    "Land",
    "Flügellänge",
    "Gewicht",
    "Tarsus",
    "Bemerkungen",
    "BeringerIn",
]

BIOMETRIC_COLS = ["Flügellänge", "Teilfederlänge", "Gewicht", "Tarsus"]
KEPT_COLS = [
    "Art",
    "Geschlecht",
    "Alter",
    "Ringstatus",
    "Fett",
    "Muskel",
    "Intensität",
    "Fortschritt",
    "Handschwingen",
    "Netz",
    "Umstand",
    "Fangmethode",
    "Lockmittel",
]


def _sample_wb():
    """A fresh, independent copy of the committed synthetic export."""
    return openpyxl.load_workbook(SAMPLE_IWM)


def _rows(ws):
    """Extract the Fangdaten data rows as ``[{header: value}, …]``."""
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    hi = {h: i for i, h in enumerate(header) if h is not None}
    out = []
    for values in it:
        if all(v is None for v in values):
            continue
        out.append({h: values[i] for h, i in hi.items()})
    return out


def _build_wb(rows, headers=CTRL_HEADERS):
    """Serialise ``rows`` (list of {header: value}) into an in-memory workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fangdaten"
    ws.append(headers)
    col = {h: i + 1 for i, h in enumerate(headers)}
    for r_i, row in enumerate(rows, start=2):
        for h, c in col.items():
            value = row.get(h)
            if value is not None:
                ws.cell(row=r_i, column=c, value=value)
    return wb


def _row(**over):
    base = {
        "Ringnummer": "V00001",
        "Ringstatus": "E",
        "Art": "Rohrammer",
        "Geschlecht": "U",
        "Alter": 3,
        "Datum": datetime(2024, 8, 1),
        "Uhrzeit": time(7, 0),
        "Ortskodierung": "NS01",
        "Geo-Koordinaten": "47.769000, 16.792000",
        "Ort": "Illmitz Zicklacke",
        "Region": "Burgenland",
        "Land": "Austria",
        "BeringerIn": "JGR",
    }
    base.update(over)
    return base


def _split_ring(ringnummer):
    match = re.match(r"^([A-Za-z]+)(\d+)$", str(ringnummer))
    assert match, f"unparseable ring number {ringnummer!r}"
    return match.group(1), match.group(2)


# --- AC: running twice on the same input produces identical output ---------


def test_running_twice_produces_identical_output():
    out1 = _rows(anonymize_workbook(_sample_wb())["Fangdaten"])
    out2 = _rows(anonymize_workbook(_sample_wb())["Fangdaten"])
    assert out1 == out2


# --- AC: collapses to exactly two Kürzel and one Station -------------------


def test_collapses_to_two_kuerzel_and_one_station():
    rows = _rows(anonymize_workbook(_sample_wb())["Fangdaten"])
    assert {r["BeringerIn"] for r in rows} == set(CURATED_KUERZEL)
    assert {r["Ort"] for r in rows} == {CURATED_STATION["name"]}
    assert {r["Ortskodierung"] for r in rows} == {CURATED_STATION["place_code"]}
    assert {r["Region"] for r in rows} == {CURATED_STATION["region"]}
    assert {r["Land"] for r in rows} == {CURATED_STATION["country"]}
    assert {r["Geo-Koordinaten"] for r in rows} == {CURATED_STATION["coordinates"]}


# --- AC: the same real Kürzel always maps to the same curated Kürzel -------


def test_same_real_kuerzel_maps_to_the_same_curated_kuerzel():
    rows = [
        _row(BeringerIn="JGR", Ringnummer="V00001"),
        _row(BeringerIn="JGR", Ringnummer="V00002"),
        _row(BeringerIn="MWA", Ringnummer="V00003"),
    ]
    out = _rows(anonymize_workbook(_build_wb(rows))["Fangdaten"])
    assert out[0]["BeringerIn"] == out[1]["BeringerIn"]
    assert out[0]["BeringerIn"] in CURATED_KUERZEL


# --- AC: ring size unchanged; numbers remapped into the demo range; injective


def test_ring_size_kept_and_numbers_remapped_injectively():
    src = _rows(_sample_wb()["Fangdaten"])
    out = _rows(anonymize_workbook(_sample_wb())["Fangdaten"])
    assert len(src) == len(out)
    mapping = {}
    for s, o in zip(src, out, strict=True):
        s_size, s_num = _split_ring(s["Ringnummer"])
        o_size, o_num = _split_ring(o["Ringnummer"])
        assert o_size == s_size  # size unchanged
        assert o_num.isdigit()
        assert int(o_num) >= DEMO_RING_BASE  # into the demo range
        assert len(o_num) >= DEMO_RING_WIDTH  # zero-padded to a fixed width
        mapping[(s_size, s_num)] = (o_size, o_num)
    # Injective: distinct real rings never collapse onto one demo ring.
    assert len(set(mapping.values())) == len(mapping)


# --- AC: Erstfang, Wiederfang and Ring-Vernichtet sharing a ring get the same


def test_shared_real_ring_gets_the_same_demo_ring():
    rows = [
        _row(Ringnummer="V00100", Ringstatus="E", Datum=datetime(2024, 7, 1)),
        _row(Ringnummer="V00100", Ringstatus="W", Datum=datetime(2024, 9, 1)),
        _row(
            Ringnummer="V00100",
            Ringstatus=None,
            Art="Ring Vernichtet",
            Datum=datetime(2024, 8, 1),
        ),
    ]
    out = _rows(anonymize_workbook(_build_wb(rows))["Fangdaten"])
    assert len({r["Ringnummer"] for r in out}) == 1


# --- AC: every date shifted by the same whole-year offset; month/day kept ---


def test_dates_shifted_by_whole_year_offset_preserving_month_and_day():
    src = _rows(_sample_wb()["Fangdaten"])
    out = _rows(anonymize_workbook(_sample_wb())["Fangdaten"])
    offset = TARGET_END_YEAR - max(r["Datum"].year for r in src)
    assert offset != 0
    for s, o in zip(src, out, strict=True):
        assert o["Datum"].year == s["Datum"].year + offset
        assert (o["Datum"].month, o["Datum"].day) == (s["Datum"].month, s["Datum"].day)


def test_wiederfang_stays_later_than_its_erstfang():
    src = _rows(_sample_wb()["Fangdaten"])
    out = _rows(anonymize_workbook(_sample_wb())["Fangdaten"])
    by_ring = defaultdict(list)
    for i, s in enumerate(src):
        by_ring[s["Ringnummer"]].append(i)
    checked = 0
    for idxs in by_ring.values():
        statuses = {src[i]["Ringstatus"] for i in idxs}
        if "E" in statuses and "W" in statuses:
            e = next(i for i in idxs if src[i]["Ringstatus"] == "E")
            w = next(i for i in idxs if src[i]["Ringstatus"] == "W")
            assert out[w]["Datum"] > out[e]["Datum"]
            checked += 1
    assert checked > 0


def test_feb29_clamps_to_feb28_in_a_nonleap_target_year():
    # max year 2020 → offset 2025-2020=5 → 29 Feb 2025 (non-leap) clamps to 28 Feb.
    rows = [_row(Ringnummer="V00001", Datum=datetime(2020, 2, 29))]
    out = _rows(anonymize_workbook(_build_wb(rows))["Fangdaten"])
    d = out[0]["Datum"]
    assert (d.year, d.month, d.day) == (TARGET_END_YEAR, 2, 28)


# --- AC: present biometrics within ±5%; absent stay blank ------------------


def test_present_biometrics_jittered_within_5pct_absent_stay_blank():
    src = _rows(_sample_wb()["Fangdaten"])
    out = _rows(anonymize_workbook(_sample_wb())["Fangdaten"])
    present = 0
    for s, o in zip(src, out, strict=True):
        for c in BIOMETRIC_COLS:
            sv, ov = s.get(c), o.get(c)
            if sv is None:
                assert ov is None
                continue
            assert ov is not None
            s_dec, o_dec = Decimal(str(sv)), Decimal(str(ov))
            assert abs(o_dec - s_dec) <= s_dec * Decimal("0.05") + Decimal("0.01")
            assert abs(Decimal(str(ov)).as_tuple().exponent) <= 2  # 2dp
            present += 1
    assert present > 0


# --- AC: Bemerkungen emptied except Aves ignota; Sonderart invariants ------


def test_bemerkungen_emptied_except_aves_ignota_placeholder():
    out = _rows(anonymize_workbook(_sample_wb())["Fangdaten"])
    aves = [r for r in out if r["Art"] == AVES_IGNOTA_ART]
    assert aves  # the sample carries Aves-ignota rows
    for r in aves:
        assert r["Bemerkungen"] == AVES_IGNOTA_PLACEHOLDER
    for r in out:
        if r["Art"] != AVES_IGNOTA_ART:
            assert not r["Bemerkungen"]


# --- AC: kept codes and species distribution unchanged ---------------------


def test_kept_fields_and_species_distribution_unchanged():
    src = _rows(_sample_wb()["Fangdaten"])
    out = _rows(anonymize_workbook(_sample_wb())["Fangdaten"])
    for s, o in zip(src, out, strict=True):
        for c in KEPT_COLS:
            assert o.get(c) == s.get(c)
    assert Counter(r["Art"] for r in src) == Counter(r["Art"] for r in out)


# --- AC: thin CLI wrapper smoke test ---------------------------------------


def test_anonymize_iwm_command_writes_safe_workbook(tmp_path):
    out_path = tmp_path / "demo_iwm.xlsx"
    call_command("anonymize_iwm", "--input", str(SAMPLE_IWM), "--output", str(out_path))
    assert out_path.exists()
    wb = openpyxl.load_workbook(out_path)
    assert "Fangdaten" in wb.sheetnames
    rows = _rows(wb["Fangdaten"])
    assert len(rows) == len(_rows(_sample_wb()["Fangdaten"]))
    assert {r["BeringerIn"] for r in rows} == set(CURATED_KUERZEL)
    assert {r["Ort"] for r in rows} == {CURATED_STATION["name"]}


def test_anonymize_iwm_command_rejects_a_non_fangdaten_file(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "SomethingElse"
    bad = tmp_path / "bad.xlsx"
    wb.save(bad)
    with pytest.raises(CommandError):
        call_command("anonymize_iwm", "--input", str(bad), "--output", str(tmp_path / "o.xlsx"))
