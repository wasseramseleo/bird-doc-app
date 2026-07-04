import zipfile
from datetime import UTC, datetime, time
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest
from django.utils.timezone import make_aware

from birds.iwm_export import SHEET_NAME, build_iwm_workbook
from birds.iwm_import import commit_import
from birds.models import Central, DataEntry, Project, Ring, RingingStation

# The one thing that differs between two builds of identical data: openpyxl
# stamps the current time into docProps/core.xml on every ``save()``. It is
# unrelated to the export's payload, so a content comparison strips it out.
_VOLATILE_MEMBERS = {"docProps/core.xml"}


def _export_payload(content):
    """The export bytes minus the live 'modified' timestamp — every meaningful
    member of the workbook zip, so two payloads compare byte-for-byte iff the
    exported content is identical."""
    src = zipfile.ZipFile(BytesIO(content))
    return {
        name: src.read(name) for name in sorted(src.namelist()) if name not in _VOLATILE_MEMBERS
    }


STATIONS_URL = "/api/birds/ringing-stations/"


def _read_all_rows(content):
    """Return a list of {header: value} dicts, one per data row, in sheet order."""
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb[SHEET_NAME]
    headers = {
        ws.cell(row=1, column=c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=c).value
    }
    rows = []
    for r in range(2, ws.max_row + 1):
        if all(ws.cell(row=r, column=col).value is None for col in headers.values()):
            continue
        rows.append({header: ws.cell(row=r, column=col).value for header, col in headers.items()})
    return rows


def _read_rows(content):
    """Return {header: value} for the first data row of the exported workbook."""
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb[SHEET_NAME]
    headers = {
        ws.cell(row=1, column=c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=c).value
    }
    return {header: ws.cell(row=2, column=col).value for header, col in headers.items()}


@pytest.mark.django_db
def test_export_emits_vienna_localtime_for_datum_and_uhrzeit(species, scientist, ringing_station):
    # 23:00 UTC is 01:00 the next day in Vienna (CEST, UTC+2). Both the date
    # and the time must reflect the Vienna wall clock the Beringer observed.
    DataEntry.objects.create(
        species=species,
        ring=Ring.objects.create(number="604", size=Ring.RingSizes.V),
        staff=scientist,
        ringing_station=ringing_station,
        date_time=datetime(2026, 6, 30, 23, 0, tzinfo=UTC),
    )

    row = _read_rows(build_iwm_workbook(DataEntry.objects.all()))

    assert row["Datum"] == datetime(2026, 7, 1, 0, 0)
    assert row["Uhrzeit"] == time(1, 0)


@pytest.mark.django_db
def test_export_emits_geschlecht_as_authentic_letters(species, scientist, ringing_station):
    # The authentic Datenmeldung carries Geschlecht as letters U/M/W, not the
    # integers our model stores (0/1/2) — this is what the importer reads back.
    for number, sex, code in (
        ("610", DataEntry.Sex.UNKNOWN, "U"),
        ("611", DataEntry.Sex.MALE, "M"),
        ("612", DataEntry.Sex.FEMALE, "W"),
    ):
        DataEntry.objects.all().delete()
        DataEntry.objects.create(
            species=species,
            ring=Ring.objects.create(number=number, size=Ring.RingSizes.V),
            staff=scientist,
            ringing_station=ringing_station,
            sex=sex,
            date_time=datetime(2026, 2, 1, 8, 0, tzinfo=UTC),
        )

        row = _read_rows(build_iwm_workbook(DataEntry.objects.all()))

        assert row["Geschlecht"] == code


@pytest.mark.django_db
def test_export_fills_land_from_station(species, scientist, ringing_station):
    ringing_station.country = "Austria"
    ringing_station.save()
    ring = Ring.objects.create(number="700", size=Ring.RingSizes.V)
    DataEntry.objects.create(
        species=species,
        ring=ring,
        staff=scientist,
        ringing_station=ringing_station,
        date_time=datetime(2026, 2, 1, 8, 0, tzinfo=UTC),
    )

    row = _read_rows(build_iwm_workbook(DataEntry.objects.all()))

    assert row["Land"] == "Austria"


@pytest.mark.django_db
def test_export_fills_region_and_ortskodierung_from_station(species, scientist, ringing_station):
    ringing_station.region = "Oberösterreich"
    ringing_station.place_code = "AU03"
    ringing_station.save()
    ring = Ring.objects.create(number="701", size=Ring.RingSizes.V)
    DataEntry.objects.create(
        species=species,
        ring=ring,
        staff=scientist,
        ringing_station=ringing_station,
        date_time=datetime(2026, 2, 1, 8, 0, tzinfo=UTC),
    )

    row = _read_rows(build_iwm_workbook(DataEntry.objects.all()))

    assert row["Region"] == "Oberösterreich"
    assert row["Ortskodierung"] == "AU03"


@pytest.mark.django_db
def test_export_geo_coordinates_as_lat_lon_with_dot_separator(species, scientist, ringing_station):
    ringing_station.latitude = Decimal("48.295892")
    ringing_station.longitude = Decimal("14.276697")
    ringing_station.save()
    ring = Ring.objects.create(number="702", size=Ring.RingSizes.V)
    DataEntry.objects.create(
        species=species,
        ring=ring,
        staff=scientist,
        ringing_station=ringing_station,
        date_time=datetime(2026, 2, 1, 8, 0, tzinfo=UTC),
    )

    row = _read_rows(build_iwm_workbook(DataEntry.objects.all()))

    assert row["Geo-Koordinaten"] == "48.295892, 14.276697"


@pytest.mark.django_db
def test_export_populates_geography_for_station_created_through_api(
    auth_client, scientist, organization, species
):
    # A Station created through the org-admin API with the required geo fields
    # carries them all the way into the IWM export (issues #114/#118).
    create = auth_client.post(
        STATIONS_URL,
        {
            "name": "Auwald Süd",
            "region": "Oberösterreich",
            "place_code": "AU03",
            "latitude": "48.295892",
            "longitude": "14.276697",
            "country": "Austria",
        },
        format="json",
    )
    assert create.status_code == 201, create.json()
    station = RingingStation.objects.get(handle=create.json()["handle"])

    DataEntry.objects.create(
        species=species,
        ring=Ring.objects.create(number="750", size=Ring.RingSizes.V),
        staff=scientist,
        ringing_station=station,
        date_time=datetime(2026, 2, 1, 8, 0, tzinfo=UTC),
    )

    row = _read_rows(build_iwm_workbook(DataEntry.objects.all()))

    assert row["Land"] == "Austria"
    assert row["Region"] == "Oberösterreich"
    assert row["Ortskodierung"] == "AU03"
    assert row["Geo-Koordinaten"] == "48.295892, 14.276697"


@pytest.mark.django_db
def test_export_fills_capture_context_from_project_defaults(
    species, scientist, ringing_station, project
):
    ring = Ring.objects.create(number="703", size=Ring.RingSizes.V)
    DataEntry.objects.create(
        species=species,
        ring=ring,
        staff=scientist,
        ringing_station=ringing_station,
        project=project,
        date_time=datetime(2026, 2, 1, 8, 0, tzinfo=UTC),
    )

    row = _read_rows(build_iwm_workbook(DataEntry.objects.all()))

    assert row["Umstand"] == "25"
    assert row["Fangmethode"] == "M"
    assert row["Lockmittel"] == "N"


@pytest.mark.django_db
def test_multi_station_project_exports_each_entrys_own_station_geography(
    species, scientist, ringing_station, organization, project
):
    ringing_station.country = "Austria"
    ringing_station.place_code = "AU03"
    ringing_station.save()
    other_station = RingingStation.objects.create(
        handle="STN2",
        name="Second Station",
        organization=organization,
        country="Germany",
        place_code="DE07",
    )
    for idx, station in enumerate((ringing_station, other_station)):
        DataEntry.objects.create(
            species=species,
            ring=Ring.objects.create(number=f"80{idx}", size=Ring.RingSizes.V),
            staff=scientist,
            ringing_station=station,
            project=project,
            date_time=datetime(2026, 2, 1, 8, idx, tzinfo=UTC),
        )

    rows = _read_all_rows(build_iwm_workbook(DataEntry.objects.order_by("date_time")))

    assert [(r["Land"], r["Ortskodierung"]) for r in rows] == [
        ("Austria", "AU03"),
        ("Germany", "DE07"),
    ]


@pytest.mark.django_db
def test_deferred_columns_remain_blank(species, scientist, ringing_station, project):
    ringing_station.country = "Austria"
    ringing_station.save()
    DataEntry.objects.create(
        species=species,
        ring=Ring.objects.create(number="900", size=Ring.RingSizes.V),
        staff=scientist,
        ringing_station=ringing_station,
        project=project,
        has_brood_patch=True,
        date_time=datetime(2026, 2, 1, 8, 0, tzinfo=UTC),
    )

    row = _read_rows(build_iwm_workbook(DataEntry.objects.all()))

    for column in ("Zustand", "Brutfleck", "Kloake"):
        assert row[column] is None


@pytest.mark.django_db
def test_export_emits_category_codes_as_text(species, scientist, ringing_station):
    # The authentic Datenmeldung carries the category fields as text codes, not
    # the raw integers our model stores — so export and import round-trip.
    DataEntry.objects.create(
        species=species,
        ring=Ring.objects.create(number="920", size=Ring.RingSizes.V),
        staff=scientist,
        ringing_station=ringing_station,
        fat_deposit=4,
        muscle_class=2,
        small_feather_int=1,
        hand_wing=3,
        net_location=7,
        date_time=datetime(2026, 2, 1, 8, 0, tzinfo=UTC),
    )

    row = _read_rows(build_iwm_workbook(DataEntry.objects.all()))

    assert row["Fett"] == "4"
    assert row["Muskel"] == "2"
    assert row["Intensität"] == "1"
    assert row["Handschwingen"] == "3"
    assert row["Netz"] == "7"


@pytest.mark.django_db
def test_export_emits_zusatzmarkierung_zz(species, scientist, ringing_station):
    # Every authentic Datenmeldung row carries Zusatzmarkierung="ZZ" (no
    # additional marking), which the importer reads back as a no-op.
    DataEntry.objects.create(
        species=species,
        ring=Ring.objects.create(number="910", size=Ring.RingSizes.V),
        staff=scientist,
        ringing_station=ringing_station,
        date_time=datetime(2026, 2, 1, 8, 0, tzinfo=UTC),
    )

    row = _read_rows(build_iwm_workbook(DataEntry.objects.all()))

    assert row["Zusatzmarkierung"] == "ZZ"


@pytest.mark.django_db
def test_sentinel_entry_exports_art_and_blank_bird_columns(
    sentinel_species, scientist, ringing_station
):
    ring = Ring.objects.create(number="601", size=Ring.RingSizes.V)
    DataEntry.objects.create(
        species=sentinel_species,
        ring=ring,
        staff=scientist,
        ringing_station=ringing_station,
        date_time=datetime(2026, 2, 1, 8, 0, tzinfo=UTC),
        bird_status=None,
        age_class=None,
        sex=None,
        comment="Produktionsfehler",
    )

    row = _read_rows(build_iwm_workbook(DataEntry.objects.all()))

    assert row["Art"] == "Ring Vernichtet"
    assert row["Ringstatus"] is None
    assert row["Alter"] is None
    assert row["Geschlecht"] is None
    assert row["Bemerkungen"] == "Produktionsfehler"


# --- Per-ring Zentrale in the Ring column (issue #230, US 20/21) ---------------
# The "Ring" column carries the ring's own EURING scheme code — never a hardcoded
# AUW. Foreign rings (a ausländischer Wiederfang) are the scientifically most
# valuable rows and were previously misreported as AUW. The "Ringnummer" column
# stays the plain Größe+Nummer concatenation — exactly what was read off the ring.


@pytest.mark.django_db
def test_domestic_entry_exports_auw_scheme_code_in_ring_column(species, scientist, ringing_station):
    # A domestic ring resolves to the AUW Zentrale (Ring.save() backfill), so the
    # Ring column stays byte-identical to today's hardcoded "AUW" output. (US 20)
    ring = Ring.objects.create(number="604", size=Ring.RingSizes.V)
    DataEntry.objects.create(
        species=species,
        ring=ring,
        staff=scientist,
        ringing_station=ringing_station,
        date_time=datetime(2026, 2, 1, 8, 0, tzinfo=UTC),
    )

    row = _read_rows(build_iwm_workbook(DataEntry.objects.all()))

    assert row["Ring"] == "AUW"
    assert row["Ringnummer"] == "V604"


@pytest.mark.django_db
def test_foreign_wiederfang_exports_its_own_scheme_code_in_ring_column(
    species, scientist, ringing_station
):
    # A ring issued by a foreign Zentrale (here the Slovak Bratislava scheme) is
    # recaptured domestically. Its Ring column must emit that ring's own scheme
    # code, not AUW. (US 20)
    skb = Central.objects.get(scheme_code="SKB")
    ring = Ring.objects.create(number="00604", size=Ring.RingSizes.V, central=skb)
    DataEntry.objects.create(
        species=species,
        ring=ring,
        staff=scientist,
        ringing_station=ringing_station,
        bird_status=DataEntry.BirdStatus.RE_CATCH,
        date_time=datetime(2026, 2, 1, 8, 0, tzinfo=UTC),
    )

    row = _read_rows(build_iwm_workbook(DataEntry.objects.all()))

    assert row["Ring"] == "SKB"


@pytest.mark.django_db
def test_foreign_ring_ringnummer_is_groesse_plus_nummer_verbatim(
    species, scientist, ringing_station
):
    # The Ringnummer cell is the free-text Größe+Nummer concatenation exactly as
    # recorded — the export writes precisely what was read off the ring, even when
    # the number carries non-numeric foreign formatting. (US 21)
    hgb = Central.objects.get(scheme_code="HGB")
    ring = Ring.objects.create(number="AB-12345", size=Ring.RingSizes.V, central=hgb)
    DataEntry.objects.create(
        species=species,
        ring=ring,
        staff=scientist,
        ringing_station=ringing_station,
        bird_status=DataEntry.BirdStatus.RE_CATCH,
        date_time=datetime(2026, 2, 1, 8, 0, tzinfo=UTC),
    )

    row = _read_rows(build_iwm_workbook(DataEntry.objects.all()))

    assert row["Ringnummer"] == "VAB-12345"


# --- Export ↔ import round-trip (issue #126) ----------------------------------
# The export now emits the same authentic codes the importer reads, so the two
# are inverses: a Projekt's captures exported to a Datenmeldung and re-imported
# reconstruct equivalent captures. Reuse of the ``build_iwm_workbook`` seam plus
# the importer's ``commit_import`` proves the whole loop, end to end.

# The capture fields the export writes and the importer reads back — the ones a
# faithful round-trip must preserve.
_ROUND_TRIP_FIELDS = (
    "species_id",
    "staff_id",
    "ringing_station_id",
    "comment",
    "sex",
    "age_class",
    "bird_status",
    "fat_deposit",
    "muscle_class",
    "small_feather_int",
    "hand_wing",
    "net_location",
    # Decimal biometrics + the moult Fortschritt (issue #176).
    "wing_span",
    "feather_span",
    "weight_gram",
    "tarsus",
    "small_feather_app",
)


@pytest.mark.django_db
def test_export_import_round_trip_reconstructs_equivalent_captures(
    species, scientist, ringing_station, project, organization
):
    ringing_station.place_code = "AU03"
    ringing_station.region = "Oberösterreich"
    ringing_station.country = "Austria"
    ringing_station.save()

    def _seed(number, **fields):
        ring, _ = Ring.objects.get_or_create(
            number=number, size=Ring.RingSizes.V, organization=organization
        )
        return DataEntry.objects.create(
            species=species,
            ring=ring,
            staff=scientist,
            ringing_station=ringing_station,
            project=project,
            organization=organization,
            **fields,
        )

    # An Erstfang carrying every round-tripping field, and its later Wiederfang
    # on the same ring (a distinct capture, not a duplicate).
    _seed(
        "00604",
        date_time=make_aware(datetime(2026, 6, 30, 8, 15)),
        bird_status=DataEntry.BirdStatus.FIRST_CATCH,
        sex=DataEntry.Sex.MALE,
        age_class=3,
        fat_deposit=4,
        muscle_class=2,
        small_feather_int=1,
        hand_wing=3,
        net_location=7,
        wing_span=Decimal("82.50"),
        feather_span=Decimal("63.00"),
        weight_gram=Decimal("18.30"),
        tarsus=Decimal("20.10"),
        small_feather_app=DataEntry.SmallFeatherAppMoult.MIXED,
        comment="Zecken am Kopf",
    )
    _seed(
        "00604",
        date_time=make_aware(datetime(2026, 7, 15, 9, 0)),
        bird_status=DataEntry.BirdStatus.RE_CATCH,
        sex=DataEntry.Sex.FEMALE,
        age_class=5,
    )

    content = build_iwm_workbook(DataEntry.objects.filter(project=project).order_by("date_time"))

    # Snapshot the source captures keyed by their capture key, then remove them so
    # the re-import is not skipped as a duplicate of the very data it round-trips.
    source = {
        (e.ring.size, e.ring.number, e.date_time): {f: getattr(e, f) for f in _ROUND_TRIP_FIELDS}
        for e in DataEntry.objects.filter(project=project)
    }
    DataEntry.objects.filter(project=project).delete()

    # A fresh Projekt in the same Organisation, so the exported Beringer Kürzel
    # and Station code resolve to the very same entities (no auto-creation).
    fresh_project = Project.objects.create(title="Round-Trip", organization=organization)
    result = commit_import(content, fresh_project)

    assert result["created"] == 2
    assert result["errors"] == []
    # The Beringer and Station already exist in the Organisation — reused, not
    # re-created — so the exported identities re-resolve exactly.
    assert result["createdBeringer"] == []
    assert result["createdStationen"] == []

    clones = {
        (e.ring.size, e.ring.number, e.date_time): e
        for e in DataEntry.objects.filter(project=fresh_project)
    }
    assert set(clones) == set(source)
    for key, want in source.items():
        got = clones[key]
        for field in _ROUND_TRIP_FIELDS:
            assert getattr(got, field) == want[field], field
        # Round-tripped into the fresh Projekt's Organisation (ADR 0005).
        assert got.organization == organization


@pytest.mark.django_db
def test_projekttyp_does_not_change_the_export(species, scientist, ringing_station, project):
    """Projekttyp is descriptive metadata, not an export field (ADR 0023): the
    IWM export is byte-for-byte identical whatever the Projekt's Projekttyp is,
    and no exported column carries it."""
    DataEntry.objects.create(
        species=species,
        ring=Ring.objects.create(number="511", size=Ring.RingSizes.V),
        staff=scientist,
        ringing_station=ringing_station,
        project=project,
        date_time=datetime(2026, 3, 1, 8, 0, tzinfo=UTC),
    )

    project.projekttyp = Project.Projekttyp.SONSTIGES
    project.save()
    baseline = build_iwm_workbook(DataEntry.objects.all())

    project.projekttyp = Project.Projekttyp.NESTLINGSBERINGUNG
    project.save()
    after = build_iwm_workbook(DataEntry.objects.all())

    # The payload (everything but the live save-timestamp) is identical.
    assert _export_payload(after) == _export_payload(baseline)
    # And there is no Projekttyp column to leak the value into.
    headers = {h for h in _read_rows(baseline) if h}
    assert not any("projekttyp" in h.lower() for h in headers)
